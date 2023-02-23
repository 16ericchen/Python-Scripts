from test_getAoi import aoi
from openpyxl import load_workbook

test = []
failed = []
failedNum = []
count = 0
failure = open("Failed Board Numbers with Errors.txt",'w')
failedNumbers = open("Failed Board Numbers.txt",'w')


projectListPath = r"K:\Testing\SPI\SPI Program List.xlsx"
wb = load_workbook(projectListPath)
ws = wb.active
for row in ws.iter_rows(min_row=2700,max_row=2950,min_col=1,max_col=7):
    if row[6].value != 'NOT YET':
        test.append(row[2].value)
for board in test:
    if aoi(board) == "True" :
        count += 1
        print(count,'/',len(test))
    else:
        failed.append(aoi(board)[1:]+[board])
        failedNum.append(board)
print(count,'/',len(test))
for d in range(len(failed)-1):
    failure.write(str(failed[d])+"\t"+'\n')
    failedNumbers.write(str(failedNum[d])+"\t"+'\n')


# board  = "920-12163-01"
# if aoi(board) == "True":
#     count += 1
#     print(count,'/',len(test))
# else:
#     failed.append(aoi(board)[1:]+[board])
#     failedNum.append(board)
# print(failed)