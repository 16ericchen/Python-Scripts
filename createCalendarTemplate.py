import os,calendar
from openpyxl import load_workbook

os.chdir(os.path.dirname(__file__))

def writeDates(dateList,month,year):
    wb = load_workbook('calendar-template.xlsx')
    if month>1:
        wb = load_workbook(year+' Calendar'+'.xlsx')
    ws = wb.worksheets[month-1]
    srow = 6
    col = 2
    for x in range(len(dateList)):
        for y in dateList[x]:
            if y>0:
                ws.cell(row = srow,column = col,value = y )
            col+=1
        col = 2
        srow+=6
    wb.save(filename=year+' Calendar'+'.xlsx')

print('Enter Calendar Year:')
year = input()
obj = calendar.Calendar()
for i in range(1,13):
    writeDates(obj.monthdayscalendar(int(year),i),i,year)



