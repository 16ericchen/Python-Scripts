from openpyxl import load_workbook,Workbook
import os
from os import path

listofBOMS = []
# status has bom as key and all the parts as value
status = {}
nums = []
mfr = {}
sortedAPN = []
rowCount = 3
info = {}
# os.chdir(path.dirname(__file__))

def enterData():
    for x in os.listdir(os.getcwd()):
        if x.endswith(('.xlsx')) and x !='BOMComparison.xlsx':
            listofBOMS.append(x)
    wb = Workbook()
    log = wb.active
    log.title = 'BOMComparison'
    for x in listofBOMS:
        parser(x)
    subsets = [[]]
    for n in list(status.keys()):
        subsets += [s + [n] for s in subsets]
    new_list = [com for com in subsets if len(com)>1]
    new_list = reversed(new_list)
    # new_list is a list of all combinations and removes all lists that have less that two values in the list and is descending in list length 
    for x in new_list:
        intersectionExport(x)
    exportHeader()

def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    start_row = 0
    mfrName,mfrPartNumber= 0,0
    status[file[:-5]] = []
    for row in sheet.iter_rows(min_row=1, min_col=1,max_col=1):
        if row[0].value is not None:
            if row[0].value.lower() == 'level':
                start_row = row[0].row
                break
    for col in sheet.iter_cols(min_row=start_row,min_col=12,max_col=15):
        if col[0].value:
            if col[0].value == 'Mfr Name':
                mfrName = col[0].column-2
            if col[0].value == 'Mfr. Part Number':
                mfrPartNumber = col[0].column-2
    for row in sheet.iter_rows(min_row = start_row+1,min_col = 2,max_col=15):
        if row[0].value:
            status[file[:-5]].append(row[0].value)
            if mfrName and mfrPartNumber:
                mfr[row[0].value] = [row[mfrName].value,row[mfrPartNumber].value]

def intersectionExport(x):
    same = status[x[0]]
    maxLen = 0
    for dataSets in x:
        maxLen = max(maxLen,len(status[dataSets]))
        same = set(same).intersection(status[dataSets])
    keyString = ','.join(list(same))
    if keyString != '':
        if keyString in info.keys():
            info[keyString][0].append([x])
        else:
            info[keyString] = [[x],maxLen]

def exportHeader():
    wb = Workbook()
    log = wb.active
    global rowCount
    log.title = 'BOMComparison'
    col = 3
    nums = sorted(info.keys(),reverse=True,key=len)
    for x in nums:
        sortedAPN.append(sorted(x.split(',')))
    for x in range(len(nums)):
        rowC = rowCount
        for combo in info[nums[x]][0]:
            log.cell(row=rowC,column=col,value=','.join(combo))
            rowC +=1
        log.cell(row=rowC,column=col,value=str(len(sortedAPN[x]))+'/'+str(info[nums[x]][1]))
        rowC+=1
        for apn in sortedAPN[x]:
            log.cell(row=rowC,column=col,value=apn)
            if apn in mfr.keys():
                log.cell(row=rowC,column=col+1,value=mfr[apn][0])
                log.cell(row=rowC,column=col+2,value=mfr[apn][1])
            rowC+=1
        exportUnique(log,info[nums[x]][0],sortedAPN[x],col+2)
    wb.save(filename='APNComparison.xlsx')  
    
def exportUnique(log,key,matches,col2):
    global rowCount
    rowC2 = rowCount
    if mfr:
        col2 +=2
    maxCurRow = 0 
    for x in key[0]:
        difference = [item for item in status[x] if item not in matches]
        log.cell(row= rowC2,column=col2,value = x)
        rowC2+=1                                                                                                             
        log.cell(row=rowC2,column = col2,value ='Unique Parts: '+str(len(difference))+'/'+(str(len(status[x]))))
        rowC2+=1
        for y in sorted(difference):
            log.cell(row=rowC2,column=col2,value = y)
            rowC2 +=1
        col2 +=3
        maxCurRow = max(maxCurRow,rowC2)
        rowC2 = rowCount 
    rowCount = maxCurRow+2
    return

enterData()