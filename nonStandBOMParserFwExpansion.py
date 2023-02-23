import os
from os import path
from openpyxl import load_workbook,Workbook
os.chdir(path.dirname(__file__))
def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    if sheet == None:
        print('Excel Format Cannot Be Read')
        print('Change File Type From Strict Open XML Spreadsheet To Regular XLSX Format')
        print('Then Restart the Program')
        input()
    dictInfo = {}
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=4):
        #0=part number,1=description,2=qty,3=ref des,4=level,5 = item Cat
        if row[0].value == None:
            continue
        if row[3].value and 'PCB' not in row[3].value:
            refTotal = row[3].value.split(',')
            for x in refTotal:
                if "-" in x:
                    print(x)
                    for y in expandArray(x):
                        dictInfo[y] = [row[0].value,row[1].value]
                #location = [APN,Description]
                else:
                    dictInfo[x] = [row[0].value,row[1].value]
    createXL(dictInfo)
def createXL(data):
    wx = Workbook()
    log = wx.active
    rnum = 1
    test = list(data.keys())
    test.sort()
    for x in test:
        log.cell(row=rnum,column=1,value= x)
        log.cell(row=rnum,column=2,value= data[x][0])
        log.cell(row=rnum,column=3,value= data[x][1])
        rnum+=1
    wx.save(filename='bom.xlsx')
def expandArray(testCase):
    #split case will be a list of the min and max range of numbers needed
    splitCase = testCase.split("-")
    if (splitCase[0][0:2]).isalpha() or splitCase[0][1] == "0":
        start = int(splitCase[0][2:])+1
        end = int(splitCase[1][2:])
        for x in range(start,end):
            splitCase.append(splitCase[0][0]+splitCase[0][1]+str(x))
    else:
        for x in range(int(splitCase[0][1:])+1,int(splitCase[1][1:])):
            splitCase.append(splitCase[0][0]+str(x))

    return splitCase
    # while cell
flag = False
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        flag = True
        parser(x)
    if x.endswith(('.xls')):
        print('Change File Type from XLS to XLSX Format')
        print('Then Restart the Program')
        input()
if flag == False:
    print("No Excel File Found")
    input()