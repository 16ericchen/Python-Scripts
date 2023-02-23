import sqlite3
from sqlite3 import Error
import os
from os import path
from openpyxl import load_workbook
import warnings
warnings.simplefilter("ignore")
apnDict = {}

# os.chdir(path.dirname(__file__))
def enterData():
    database = r"N:\Python\Project Database\APN.db"
    conn = create_connection(database)
    cur = conn.cursor()
    cur.execute("SELECT ALL APN FROM projects")
    y = cur.fetchall()
    for z in y:
        if z not in apnDict.keys():
            cur.execute("DELETE FROM projects WHERE (APN) = (?)",(z[0],))
            conn.commit()
    for x in apnDict:
        cur.execute("SELECT * FROM projects WHERE (APN) = (?)",(x,))
        c = cur.fetchall()
        projects = ','.join(apnDict[x][0])
        quantity = ','.join(apnDict[x][1])
        task_1 = x,projects,quantity
        if c:
            if projects == c[0][0] or quantity == c[0][1]:
                continue
            else:
                update_projects(conn,task_1)
        else:
            create_projects(conn,task_1)
    print("Done")

def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except Error as e:
        print(e)
    return conn

def create_table(conn, create_table_sql):
    try:
        c = conn.cursor()
        c.execute(create_table_sql)
    except Error as e:
        print(e)

def create_projects(conn, projects):
    sql = ''' INSERT INTO projects(APN,Project_Name,QTY)
            VALUES(?,?,?) '''
    cur = conn.cursor()
    cur.execute(sql, projects)
    conn.commit()
    return cur.lastrowid

def update_projects(conn,projects):
    sql = ''' UPDATE projects SET Project_Name = ?,QTY = ? WHERE APN = ?'''
    cur = conn.cursor()
    # need to make sure that the list is in the same order as the sql statement
    cur.execute(sql,[projects[1],projects[2],projects[0]])
    conn.commit()
    return cur.lastrowid

def createTable():
    database = r"N:\Python\Project Database\APN.db"
    conn = create_connection(database)
    sql_create_projects_table = """ CREATE TABLE IF NOT EXISTS projects (
                                            APN text PRIMARY KEY,
                                            Project_Name text,
                                            QTY text
                                        ); """
    create_table(conn, sql_create_projects_table)

def subFolders(path,names):
    for x in names:
        for y in os.listdir(path+'\\'+x):
            if y.endswith(('.xlsx')) and "~" not in y and 'UPDATED' in y:
                parseXL(path+'\\'+x+'\\'+y,x) 
                
def parseXL(file,name):
    wb = load_workbook(filename=file,data_only=True)
    sheet = wb.active
    maxCol = sheet.max_column-1
    while sheet.cell(row=1,column=maxCol).value != 'O/H AS UPDATED':
        maxCol -= 1
    for row in sheet.iter_rows(min_row=3, min_col = 1,max_col=sheet.max_column):
        if row[0].value not in apnDict.keys():
            apnDict[row[0].value] = [[name],[str(row[maxCol-1].value)]]
        else:
            apnDict[row[0].value][0].append(name)
            apnDict[row[0].value][1].append(str(row[maxCol-1].value))

def getData():
    path = r"N:\WHS\Project Inventory"
    folders = next(os.walk(path))[1]
    for x in folders:
        sub = next(os.walk(path+"\\"+x))[1]
        if 'OLD' in sub:
            sub.remove('OLD')
        if '.DAV' in sub:
            sub.remove('.DAV')
        if sub:
            subFolders(path+"\\"+x,sub)
        else:
            for y in os.listdir(path+'\\'+x):
                if y.endswith(('.xlsx')) and "~" not in y and 'UPDATED' in y:
                    parseXL(path+'\\'+x+'\\'+y,x)
def main():
    database = r"N:\Python\Project Database\APN.db"
    getData()
    conn = create_connection(database)
    createTable()
    if conn is not None:
        enterData()
    else:
        print("Error! cannot create the database connection.")

if __name__ == '__main__':
    main()

