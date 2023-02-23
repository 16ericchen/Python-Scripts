from os.path import exists
import os
from openpyxl import load_workbook,Workbook
from collections import OrderedDict
from openpyxl.styles import Font
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
os.chdir(os.path.dirname(__file__))
#probably don't need these many variables so maybe optimize the variable usage later on
x,y,z,Mismatch,partnumber,noLoad = [],[],[],[],[],[]
top,bot,noStuffTop,noStuffBot = [],[],[],[]
dict1,dict2 = {},{}
topd,botd,noStuffTopd,noStuffBotd ={},{},{},{}
topDesQT,botDesQT,noStuffTopdDesQT,noStuffBotdDesQT = {},{},{},{}
#changes the current directory to where the python file is
#reason is that there were issues where the directory was pointed to a different location

# Function that takes two variables side = an input list and tf = filename that the input list is being written to
def writetotextfile(side,tf):
        file = open(tf,'w')
        for d in side:
            for x in d:
                file.write(str(x)+"\t")
            file.write('\n')   

# Function that reads the comp.txt file and formats the information and appends them to a dictionary
def readCompFile(inputArray):
    for line in inputArray:
        if '"' in line:
            x.append(line.strip().split('"'))
        else:
            x.append(line.strip().split(','))
    for item in x[5:]:
        # checks to see if the string is already seperated correctly 
        if item[-1] == 'NO' or item[-1] == 'YES':
            y.append(item[0:2]+item[-1:])
            dict1[item[0]] = item[-1]
        # if the string is not seperated properly already its will split it again
        else:
            y.append([item[0].strip(',')]+item[1:2]+[item[-1][-3:].strip(',')])
            dict1[item[0].strip(',')] = item[-1][-3:].strip(',')
    inputArray.close()
    x.clear()
    y.sort()
    

def readBomFile():
    wb = load_workbook(filename="bom.xlsx")
    sheet = wb.active
    for row in sheet.iter_rows(min_row = 1,max_row = None,min_col=1,max_col=3):
        dict2[str(row[0].value)] = ([str(row[1].value),str(row[2].value)])
        partnumber.append(str(row[1].value))

#  if ref des in Comp but not in bom == no stuff
#  if ref des not in comp but in bom == mismatch
#  if ref des in comp and bom == output

def createAOI():
    for i in range(len(y)):
        if y[i][0] in dict2:
            side(y[i][2],topd,botd,[y[i][0]]+dict2[y[i][0]][1:2],dict2[y[i][0]][0])
        else:
            noLoad.append(y[i][0])
            side(y[i][2],noStuffTopd,noStuffBotd,y[i][:-1],'No Load')
    for x in dict2.keys():
        if x not in dict1:
            Mismatch.append([x,dict2[x][0],dict2[x][1]])
    if Mismatch:
        writetotextfile(Mismatch,"Mismatch and FM.txt")
    topdE = OrderedDict(sorted(topd.items()))
    botdE = OrderedDict(sorted(botd.items()))
    noStuffBotdE = OrderedDict(sorted(noStuffBotd.items()))
    noStuffTopdE = OrderedDict(sorted(noStuffTopd.items()))
    exportCAD(y,'cad.xlsx')
    exportExcel(topd,'tload.xlsx')
    exportExcel(botd,'bload.xlsx')
    exportExcelNo(noStuffTopd,'tnoload.xlsx')
    exportExcelNo(noStuffBotd,'bnoload.xlsx')
    shorten(topdE,topDesQT)
    shorten(botdE,botDesQT)
    shorten(noStuffTopdE,noStuffTopdDesQT)
    shorten(noStuffBotdE,noStuffBotdDesQT)
    exportMPIXL(topdE,botdE,noStuffTopdE,noStuffBotdE,topDesQT,botDesQT,noStuffTopdDesQT,noStuffBotdDesQT)

def side(TB,topDict,botDict,value,key):
    if TB == 'NO':
        if key in topDict.keys():
            topDict[key].append(value+['( T )'])
        else:
            topDict[key] = []
            topDict[key].append(value+['( T )'])
    if TB == 'YES':
        if key in botDict.keys():
            botDict[key].append(value+['( B )'])
        else:
            botDict[key] = []
            botDict[key].append(value+['( B )'])

def shorten(inDict,outDict):
    for x in inDict:
        min = inDict[x][0][0]
        max = inDict[x][0][0]
        loopCount = 0
        hyphenCount = 0
        newList = []
        outDict[x] = [inDict[x][0][1],inDict[x][0][2],len(inDict[x])]
        if x == 'No Load':
            outDict[x] = ['No Load',inDict[x][0][2],len(inDict[x])]
        while loopCount<len(inDict[x]):
            if loopCount+1 == len(inDict[x]):
                if hyphenCount > 1:
                    newList.append(min+'-'+max)
                    loopCount+=1
                else:
                    if min != max:
                        newList.append(min)
                        newList.append(max)
                        loopCount += 1
                    else:
                        newList.append(min)
                        loopCount += 1
            else:
                if addOne(max) == inDict[x][loopCount+1][0]:
                    loopCount += 1
                    hyphenCount += 1
                    max = inDict[x][loopCount][0]
                else:
                    if hyphenCount > 1:
                        newList.append(min+'-'+max)
                        loopCount+=1
                        hyphenCount = 0
                        min = inDict[x][loopCount][0]
                        max = inDict[x][loopCount][0]

                    else:
                        if min != max:
                            newList.append(min)
                            newList.append(max)
                            hyphenCount = 0
                            loopCount += 1
                            min = inDict[x][loopCount][0] 
                            max = inDict[x][loopCount][0]

                        else:
                            newList.append(min)
                            hyphenCount = 0
                            loopCount += 1
                            min = inDict[x][loopCount][0]
                            max = inDict[x][loopCount][0]
        inDict[x] = newList
def createFiles(inputFile):
    f = open(inputFile,"r",encoding='utf-8')
    readCompFile(f)
    readBomFile()
    createAOI()

def addOne(num):
    carry = 0
    output = ''
    for x in reversed(range(len(num))):
        if num[x].isdigit():
            checkNum = int(num[x])+carry
            if x == len(num)-1:
                checkNum = int(num[x])+carry+1
            if checkNum < 10:
                output = str(checkNum) + output
                carry = 0
            else:
                output = str(checkNum-10) + output
                carry = 1
        else:
            if carry == 1:
                output = num[x] + '1' + output
                carry = 0
            else:
                output = num[x]+output
    return output

def exportCAD(data,fileName):
    wb = Workbook()
    log = wb.active
    log.title = fileName[:-5]
    for i,x in enumerate(data):
        log.cell(row=i+1,column=1,value=x[0])
        log.cell(row=i+1,column=2,value=x[1])
        log.cell(row=i+1,column=3,value=x[2])
    wb.save(filename=fileName)

def exportExcel(data,fileName):
    wb = Workbook()
    log = wb.active
    log.title = fileName[:-5]
    columnWidths = [11.14,11.14,20,11.14]
    rnum = 0
    cellFont = Font(name = 'Times New Roman',size = 11)
    columnL = ['A','B','C','D']
    for i in range(4):
        log.column_dimensions[columnL[i]].width = columnWidths[i]
    for x in data:
        for y in range(len(data[x])): #iterate row, iterates through all the keys in the dictionary,
            rnum +=1
            cnum = 1
            for z in range(len(data[x][y])): #iterate column,iterates through the values associated with the dictionary key,dict1[x][y][z] == item in the list(y) that is associated with the key(x)
                if z == 1:
                    dataCell = log.cell(row = rnum,column=cnum+z,value = x)
                    dataCell.font = cellFont
                    cnum +=1
                    dataCell = log.cell(row = rnum,column=cnum+z,value = data[x][y][z])
                    dataCell.font = cellFont
                else:
                    dataCell = log.cell(row = rnum,column=cnum+z,value = data[x][y][z])
                    dataCell.font = cellFont
    wb.save(filename=fileName)
    return

def exportExcelNo(data,fileName):
    wb = Workbook()
    log = wb.active
    log.title = 'MPI'
    columnWidths = [11.14,11.14,20,11.14]
    rnum = 0
    cellFont = Font(name = 'Times New Roman',size = 11)
    columnL = ['A','B','C','D']
    for i in range(4):
        log.column_dimensions[columnL[i]].width = columnWidths[i]
    for x in data:
        for y in range(len(data[x])): #iterate row, iterates through all the keys in the dictionary,
            rnum +=1
            cnum = 1
            for z in range(len(data[x][y])+3):#iterate column,iterates through the values associated with the dictionary key,dict1[x][y][z] == item in the list(y) that is associated with the key(x)
                if z+cnum < 3:
                    dataCell = log.cell(row = rnum,column=cnum+z,value = "No Load")
                    dataCell.font = cellFont
                if z+cnum == 3:
                    continue
                if z+cnum > 3:
                    dataCell = log.cell(row = rnum,column=cnum+z,value = data[x][y][z-3])
                    dataCell.font = cellFont
    wb.save(filename=fileName)
    return

def exportMPIXL(top,bot,noTop,noBot,top2,bot2,noTop2,noBot2):
    wb = Workbook()
    log = wb.active
    log.title = 'MPI'
    borderStyle = borders.Side(style= None, color='FF000000',border_style='thin' )
    thin = Border(left=borderStyle,right=borderStyle,bottom=borderStyle,top=borderStyle)
    rnum = 1
    cnum = 1
    columnWidths = [10.38,45.63,4.63,47.75,22.63] #for mpi width
    columnL = ['A','B','C','D','E']
    for i in range(5):
        log.column_dimensions[columnL[i]].width = columnWidths[i]
    len1 = workbookPass(log,bot,bot2,cnum,rnum,'SMT Bottom Side Load')
    len2 = workbookPass(log,noBot,noBot2,cnum,len1+2,"SMT Bottom Side No Load")
    len3 = workbookPass(log,top,top2,cnum,len2+2,"SMT Top Side Load")
    len4 = workbookPass(log,noTop,noTop2,cnum,len3+2,'SMT Top Side No Load')
    for row in log.iter_rows(min_row=1, min_col=1, max_row=len4+1, max_col=5):
        for cell in row:
            cell.border = thin
    log.oddFooter.center.text = "Page &P of &N"
    # log.evenFooter.center.text = "Page &P of &N"   
    log.oddHeader.center.text = "Project Name"
    log.oddHeader.center.size = 12
    log.oddHeader.center.font = "Times New Roman"
    log.oddFooter.center.size = 12
    log.oddFooter.center.font = 'Times New Roman'
    log.oddFooter.left.size = 12
    log.oddFooter.left.font = 'Times New Roman'
    log.oddFooter.left.text = ''
    log.page_margins.left = 0.25
    log.page_margins.right = 0.2
    log.page_margins.top = 0.57
    log.page_margins.bottom = 0.47
    log.page_margins.header = 0.26
    log.page_margins.footer = 0.17
    log.page_setup.orientation = log.ORIENTATION_LANDSCAPE
    wb.save(filename='mpi.xlsx')


def workbookPass(work,info,info2,cnum,rnum,title):
    titleFont = Font(name = 'Times New Roman',size = 11,bold=True)
    cellFont = Font(name = 'Times New Roman',size = 11)
    dataCell = work.cell(row=rnum,column=cnum,value = title)
    final = 0
    dataCell.font = titleFont
    if info2:
        for x in info:
            test = '' 
            dataCell = work.cell(row=rnum+1,column=cnum,value= x)#part number
            dataCell.font = cellFont
            dataCell = work.cell(row=rnum+1,column=cnum+1,value= info2[x][0])#specs
            dataCell.font = cellFont
            dataCell = work.cell(row=rnum+1,column=cnum+2,value= info2[x][2])#how many
            dataCell.font = cellFont
            final += info2[x][2]#len(info[x])
            if len(info[x])>1:
                test = info[x][0]
                for y in range(1,len(info[x])):
                    if(len(test)+len(info[x][y])+1>50):
                        dataCell = work.cell(row=rnum+1,column = cnum+3,value = test)
                        dataCell.font = cellFont
                        test = info[x][y]
                        rnum+=1
                    else:
                        test+= ","+info[x][y]
                dataCell=work.cell(row=rnum+1,column = cnum+3,value=test)
                dataCell.font = cellFont
            else:
                dataCell=work.cell(row=rnum+1,column = cnum+3,value=info[x][0])
                dataCell.font = cellFont
            rnum+=1
        if "No Load" not in info.keys():
            # print(info.keys())
            dataCell = work.cell(row = rnum+1,column = 3,value = final)
        dataCell.font = cellFont
    return rnum

if exists("comp.txt") and exists("bom.xlsx"):
    createFiles("comp.txt")
else:
    print("No Comp.txt and BOM.xlsx found")
    input()