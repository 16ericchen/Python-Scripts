# Need to get list of part numbers already in the MM
# Need to get list of part numbers in the new Revision BOM
# MM will always be the first sheet and the second Sheet will always be the new Revision BOM
# Also check and update the quantity if it needs to be updated
# If there is something that is in the old revision but not in the new BOM delete row and move everything up
# Then place the deleted information at the bottom of the list or a few after the list
#information to grab from the BOM is Part Number, Description, QTY,MFR, and MPN


from os import path
from openpyxl import load_workbook
import os
from openpyxl.styles import Font

os.chdir(path.dirname(__file__))
mm = {}
bom = {}
def parser(file):
    wb = load_workbook(filename=file)
    sheets = wb.sheetnames
    getBOMData(wb[sheets[1]])
    getMMData(wb[sheets[0]])
    wb.save(filename=file)

def getMMData(sheet):
    bom.pop(None)
    rows = list(sheet.iter_rows(min_row=4, max_row=sheet.max_row))
    rows = reversed(rows)
    for row in rows:
        if row[1].value in bom.keys():
            if row[3].value != bom[row[1].value][3]:
                row[3].value = bom[row[1].value][3]
                row[3].font = Font(bold=True)
            bom.pop(row[1].value)
        else:
            print(row[1].value)
            print('delete')

def getBOMData(sheet):
    for row in sheet.iter_rows(min_row = 2, min_col = 1):
        bom[row[2].value] = []
        for i in range(len(row)):
            if i != 2:
                bom[row[2].value].append(row[i].value)
            else:
                continue
    # print(bom)

for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        flag = True
        parser(x)

if flag == False:
    print("No Excel File Found")
    input()