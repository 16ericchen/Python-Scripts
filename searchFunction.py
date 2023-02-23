from openpyxl import load_workbook,Workbook
import os
from os import path
listofBOMS = []
status = {}
# os.chdir(path.dirname(__file__))
def enterData(APN):
    for x in os.listdir(os.getcwd()):
        if x.endswith(('.xlsx')) and x !='BOMComparison.xlsx':
            listofBOMS.append(x)
    wb = Workbook()
    log = wb.active
    log.title = 'BOMComparison'
    for x in listofBOMS:
        parser(x,APN)
    return status

def parser(file,apn):
    wb = load_workbook(filename=file)
    sheet = wb.active
    start_row = 0
    for row in sheet.iter_rows(min_row=1, min_col=1,max_col=1):
        if row[0].value is not None:
            if row[0].value.lower() == 'level':
                start_row = row[0].row
                break
    for row in sheet.iter_rows(min_row = start_row,min_col = 2,max_col=3):
        if row[0].value == apn or row[1].value == apn:
            status[file[:-5]] = 'APN Found'
            return
    status[file[:-5]] = 'APN Not Found'
    return
    