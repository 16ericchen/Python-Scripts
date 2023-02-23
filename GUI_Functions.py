import sqlite3,datetime
from sqlite3 import Error
from consolemenu import *
from consolemenu.items import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font
from openpyxl.utils import get_column_letter

database = r"C:\Users\EricChen\Desktop\Shipping Database\testDatabase.db"
time = datetime.date.today()
projectDate = time.strftime("%m-%d-%Y")
def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except Error as e:
        print(e)
    return conn

def create_table(conn, create_table_sql):
    try:
        c = conn.cursor()
        c.execute(create_table_sql)
    except Error as e:
        print(e)

def exportExcel():
    global database
    wb = Workbook()
    dest_filename = 'daily_shipping_log'+"("+projectDate+")"+'.xlsx'
    log = wb.active
    log.title = projectDate
    conn = create_connection(database)
    cur = conn.cursor()
    cur.execute("SELECT * FROM projects WHERE date=?", (projectDate,))
    x = cur.fetchall()
    inputItems = ["Project","Assembly","Serial Number","Matrix","Note"]
    columnWidths = [18.5,22,33.5,17,12]
    createTemplate(log,len(x),inputItems)
    for y in range(0,(len(x)*5),5):
        cellFont = Font(name = 'Calibri',size = 18)
        projectCell = log.cell(column = y+1, row = 2, value = x[y//5][0])
        assemblyCell = log.cell(column = y+2, row = 2, value = x[y//5][1])
        log.column_dimensions[get_column_letter(y+1)].width = columnWidths[y%5]
        log.column_dimensions[get_column_letter(y+2)].width = columnWidths[(y+1)%5]
        projectCell.font = cellFont
        assemblyCell.font = cellFont
        cur.execute("SELECT serialNumber,matrix,note FROM serialTable WHERE (project_name,assembly_name,date) = (?,?,?)",(x[y//5][0],x[y//5][1],x[y//5][2]))
        i = cur.fetchall()
        addSerial(y,i,log)
    wb.save(filename=dest_filename)
    return

def addSerial(columnNumber,values,sheet):
    cellFont = Font(name = 'Calibri',size = 18)
    columnWidths = [33.5,17,12]
    for x in range(2,len(values)+2):
        cell1 = sheet.cell(column = columnNumber+3, row = x, value = values[x-2][0])
        cell2 = sheet.cell(column = columnNumber+4, row = x, value = values[x-2][1])
        cell3 = sheet.cell(column = columnNumber+5, row = x, value = values[x-2][2])
        sheet.column_dimensions[get_column_letter(columnNumber+3)].width = columnWidths[0]
        sheet.column_dimensions[get_column_letter(columnNumber+4)].width = columnWidths[1]
        sheet.column_dimensions[get_column_letter(columnNumber+5)].width = columnWidths[2]
        cell1.font = cellFont
        cell2.font = cellFont
        cell3.font = cellFont

def createTemplate(workbook,amount,inputItems):
    i = 0 
    labelFonts = Font(name = 'Verdana',size = 18,bold=True)
    labelFill1 = PatternFill('solid','FFFF00')
    labelFill2 = PatternFill('solid','E26B0A')
    fillLabel = labelFill1
    for x in range(1,amount*5+1):
        test = workbook.cell(column = x ,row = 1,value = inputItems[i])
        test.fill = fillLabel
        test.font = labelFonts
        if i == 4:
            if (x % 2) == 0:
                fillLabel = labelFill1
            else:
                fillLabel = labelFill2
            i = 0
        else:
            i+=1

def create_projects(conn, projects):
    sql = ''' INSERT INTO projects(name,assembly,date)
              VALUES(?,?,?) '''
    cur = conn.cursor()
    cur.execute("SELECT * FROM projects WHERE (name,assembly,date) = (?,?,?)",(projects[0],projects[1],projects[5],))
    x = cur.fetchall()
    if not x:
        cur.execute(sql, [projects[0],projects[1],projects[5]])
    conn.commit()
    return cur.lastrowid

def create_serial(conn, projects):
    sql = ''' INSERT INTO serialTable(project_name,assembly_name,matrix,note,serialNumber,date)
              VALUES(?,?,?,?,?,?) '''
    cur = conn.cursor()
    cur.execute("SELECT * FROM serialTable WHERE serialNumber=?", (projects[4], ))
    x =cur.fetchone()
    if x:
        return x
    else:
        cur.execute(sql, projects)
    conn.commit()
    return False

def deleteEntry(project):
    global database
    conn = create_connection(database)
    cur = conn.cursor()
    cur.execute('DELETE FROM serialTable WHERE serialNumber=?', (project[4],))
    conn.commit()

def createTable():
    global database
    conn = create_connection(database)
    sql_create_projects_table = """ CREATE TABLE IF NOT EXISTS projects (
                                            name text NOT NULL,
                                            assembly text NOT NULL,
                                            date text
                                        ); """
    sql_create_serialNumber_table = """CREATE TABLE IF NOT EXISTS serialTable (
                                    project_name text NOT NULL,
                                    assembly_name text,
                                    matrix text,
                                    note text,
                                    serialNumber CHAR(25) PRIMARY KEY,
                                    date text,
                                    FOREIGN KEY (project_name) REFERENCES projects (name),
                                    FOREIGN KEY (assembly_name) REFERENCES projects (assembly)
                                );"""
    create_table(conn, sql_create_projects_table)
    create_table(conn, sql_create_serialNumber_table)

def enterData(project):
        createTable()
        global database
        conn = create_connection(database)
        if conn is not None:
            project.append(projectDate)
            create_projects(conn,project)
            if create_serial(conn,project):
                return create_serial(conn,project)
            else:
                return False
        else:
            print("Error! cannot create the database connection.")

        

