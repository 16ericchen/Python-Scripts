from openpyxl import load_workbook,Workbook
import os
from collections import OrderedDict
os.chdir(os.path.dirname(__file__))
i = 2
listofBOMS = []
component = {}
masterList = {}
cur = ''
newKey = ''


def parser(file,count):
    wb = load_workbook(filename=file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, min_col=1):
        if row:
            print(row[0].value)
            if row[0].value in component.keys():
                if len(component[row[0].value]) == count:
                    component[row[0].value] = component[row[0].value]+[row[1].value]
                else:
                    fill = ['No Stuff']*(count-len(component[row[0].value]))
                    component[row[0].value] = component[row[0].value]+fill+[row[1].value]
            else:
                if count>0:
                    fill = ['No Stuff']*count
                    component[row[0].value] = fill + [row[1].value]
                else:
                    component[row[0].value] = [row[1].value]
        else:
            break
    

for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')) and x !='BOMComparison.xlsx':
        listofBOMS.append(x)
wb = Workbook()
log = wb.active
log.title = 'BOMComparison'
for x in range(len(listofBOMS)):
    log.cell(row=1,column=x+4,value=listofBOMS[x][-13:-5])
    parser(listofBOMS[x],x)
final= OrderedDict(sorted(component.items()))
for x,y in final.items():
    if len(y)<len(listofBOMS):
        fill = ['No Stuff']*(len(listofBOMS)-len(y))
        y = y + fill
    if len(set(y))>1:
        log.cell(row=i,column=3,value=x)
        for z in range(len(y)):
            log.cell(row=i,column=z+4,value=y[z])
        i +=1
wb.save(filename='BOMComparison.xlsx')       