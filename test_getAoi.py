import os
from openpyxl import load_workbook
def aoi(boardNum):
    desktop = os.path.join(os.environ['USERPROFILE'],'Desktop','AOI')  
    projectListPath = r"K:\Testing\SPI\SPI Program List.xlsx"
    simpleName = projectInfo = fullName =  ''
    wb = load_workbook(projectListPath)
    ws = wb.active
    for row in ws.iter_rows(min_row=2700,max_row=2950,min_col=1,max_col=7):
        if (boardNum == row[2].value) and (row[6].value == 'NOT YET'):
            return 'True'
        if (boardNum == row[2].value) and (row[6].value!='NOT YET'):
            projectInfo = row[0].value
    if projectInfo == '':
        return ["False","Board Number Not Found"]
    projectInfo = projectInfo.replace(',','')
    if not os.path.exists(desktop):
        os.makedirs(desktop)
    folderPath = os.path.join(desktop,projectInfo)
    if not os.path.exists(folderPath):
        os.makedirs(folderPath)
    projectLetter = projectInfo[0]
    test = projectInfo.split(' ')
    test = [i for i in test if i]
    projectName = ' '.join(test[0:-2])
    if 'board' in projectInfo:
        simpleName = projectInfo.split('board')[0]
    if 'board' not in projectInfo and simpleName == '':
        simpleName = projectInfo.split('Proto')[0]
    projectNumber = test[0]
    projectDate = test[-1]
    projectFolderPath = "K:\Projects"+"\\"+projectLetter+"\\"+projectNumber
    if not os.path.exists(projectFolderPath):
        return ["False",'Folder Path Does Not Exists',projectFolderPath]
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if x.replace(' ','').lower() in projectName.replace(' ','').lower() or simpleName.replace(' ','').lower() in x.replace(' ','').lower():
            fullName =projectFolderPath+ "\\"+x
            if simpleName.replace(' ','').lower() == x.replace(' ','').lower():
                break
    projectFolderPath = fullName
    if not os.path.exists(projectFolderPath):
        return ["False",'Folder Path Does Not Exists',projectFolderPath]
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if (x[:-11] in projectName or projectDate==folders[0][-10:]) and x[:-11] != '':
            if x[:-11] == projectName:
                projectFolderPath = projectFolderPath +"\\"+x
                break
    if not os.path.exists(projectFolderPath):
        return ["False",'Folder Path Does Not Exists',projectFolderPath]
    compPath = projectFolderPath + '\\'+'Gerber'+'\\'+'comp.txt'
    if not os.path.exists(compPath):
        return ["False",'COMP FILE Does Not Exists',projectFolderPath]
    bomPath = projectFolderPath + '\\'+'BOM_CAD'
    if not os.path.exists(bomPath):
        return ["False",'BOM File Does Not Exists',bomPath]
    buildMatrixPath = projectFolderPath +'\\'+'Information'
    if not os.path.exists(buildMatrixPath):
        return ["False",'Folder Path Does Not Exists',buildMatrixPath]
    return "True"

