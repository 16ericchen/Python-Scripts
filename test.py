from os.path import exists
from runpy import run_module
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill,Font
from openpyxl.utils import get_column_letter

x,y,z,Mismatch = [],[],[],[]
top,bot,noStuffTop,noStuffBot = [],[],[],[]
dict1,dict2 = {},{}
topd,botd,noStuffTopd,noStuffBotd ={},{},{},{}

def writetotextfile(side,tf):
        for d in side:
            tf.write(str(d)+"\t")
        tf.write('\n')   

def readCompFile(inputArray):
    for line in inputArray:
        if '"' in line:
            x.append(line.strip().split('"'))
        else:
            x.append(line.strip().split(','))
    for item in x[5:]:
        if item[-1] == 'NO' or item[-1]== 'Yes':
            y.append(item[0:2]+item[-1:])
            dict1[item[0]] = item[-1]
        else:
            y.append([item[0].strip(',')]+item[1:2]+[item[-1][-3:].strip(',')])
            dict1[item[0].strip(',')] = item[-1][-3:].strip(',')
    inputArray.close()
    x.clear()
    y.sort()

def readBomFile():
    wb = load_workbook(filename="bom.xlsx")
    sheet = wb.active
    for row in sheet.iter_rows(min_row = 1,max_row = None,min_col=1,max_col=3):
        dict2[str(row[0].value)] = ([str(row[1].value),str(row[2].value)])

#  if ref des in Comp but not in bom == no stuff
#  if ref des not in comp but in bom == mismatch
#  if ref des in comp and bom == output

def createAOI():
    for i in range(len(y)):
        if y[i][0] in dict2:
            side(y[i][2],topd,botd,[y[i][0]]+dict2[y[i][0]][1:2],dict2[y[i][0]][0])
        else:
            side(y[i][2],noStuffTopd,noStuffBotd,y[i][:-1],'No Load')
    for x in dict2.keys():
        if x[0] not in dict1:
            Mismatch.append(x)
    # if Mismatch:
    #     writetotextfile(Mismatch)
    exportExcel(topd,'tload.xlsx')
    exportExcel(botd,'bload.xlsx')
    exportExcelNo(noStuffTopd,'tnoload.xlsx')
    exportExcel(noStuffBotd,'bnoload.xlsx')
    exportMPIXL(topd,botd,noStuffTopd,noStuffBotd)

def side(TB,topDict,botDict,value,key):
    if TB == 'NO':
        if key in topDict.keys():
            topDict[key].append(value+['( T )'])
        else:
            topDict[key] = []
            topDict[key].append(value+['( T )'])
    if TB == 'YES':
        if key in botDict.keys():
            botDict[key].append(value+['( B )'])
        else:
            botDict[key] = []
            botDict[key].append(value+['( B )'])

def createFiles(inputFile):
    f = open(inputFile,"r",encoding='utf-8')
    readCompFile(f)
    readBomFile()
    createAOI()

def exportExcel(data,fileName):
    wb = Workbook()
    log = wb.active
    log.title = 'MPI'
    columnWidths = [11.14,11.14,20,11.14]
    rnum = 0
    cellFont = Font(name = 'Times New Roman',size = 11)
    columnL = ['A','B','C','D']
    for i in range(4):
        log.column_dimensions[columnL[i]].width = columnWidths[i]
    for x in data:
        for y in range(len(data[x])): #iterate row, iterates through all the keys in the dictionary,
            rnum +=1
            cnum = 1
            for z in range(len(data[x][y])): #iterate column,iterates through the values associated with the dictionary key,dict1[x][y][z] == item in the list(y) that is associated with the key(x)
                if z == 1:
                    dataCell = log.cell(row = rnum,column=cnum+z,value = x)
                    dataCell.font = cellFont
                    cnum +=1
                    dataCell = log.cell(row = rnum,column=cnum+z,value = data[x][y][z])
                    dataCell.font = cellFont
                else:
                    dataCell = log.cell(row = rnum,column=cnum+z,value = data[x][y][z])
                    dataCell.font = cellFont
    wb.save(filename=fileName)
    return

def exportExcelNo(data,fileName):
    wb = Workbook()
    log = wb.active
    log.title = 'MPI'
    columnWidths = [11.14,11.14,20,11.14]
    rnum = 0
    cellFont = Font(name = 'Times New Roman',size = 11)
    columnL = ['A','B','C','D']
    for i in range(4):
        log.column_dimensions[columnL[i]].width = columnWidths[i]
    for x in data:
        for y in range(len(data[x])): #iterate row, iterates through all the keys in the dictionary,
            rnum +=1
            cnum = 1
            for z in range(len(data[x][y])+3):#iterate column,iterates through the values associated with the dictionary key,dict1[x][y][z] == item in the list(y) that is associated with the key(x)
                if z+cnum < 3:
                    dataCell = log.cell(row = rnum,column=cnum+z,value = "No Load")
                    dataCell.font = cellFont
                if z+cnum == 3:
                    continue
                if z+cnum > 3:
                    dataCell = log.cell(row = rnum,column=cnum+z,value = data[x][y][z-3])
                    dataCell.font = cellFont
    wb.save(filename=fileName)
    return

def exportMPIXL(top,bot,noTop,noBot):

    wb = Workbook()
    log = wb.active
    log.title = 'MPI'
    rnum = 1
    cnum = 1
    columnWidths = [10.38,45.63,4.63,47.75] #for mpi width
    columnL = ['A','B','C','D']
    for i in range(4):
        log.column_dimensions[columnL[i]].width = columnWidths[i]
    workbookPass(log,bot,cnum,rnum,'SMT Bottom Side Load')
    workbookPass(log,noBot,cnum,len(bot)+4,"SMT Bottom Side No Load")
    workbookPass(log,top,cnum,len(bot)+len(noBot)+7,"SMT Top Side Load")
    workbookPass(log,noTop,cnum,len(bot)+len(noBot)+len(top)+10,'SMT Top Side No Load')
    wb.save(filename='mpi.xlsx')


def workbookPass(work,info,cnum,rnum,title):
    titleFont = Font(name = 'Times New Roman',size = 11,bold=True)
    cellFont = Font(name = 'Times New Roman',size = 11)
    dataCell = work.cell(row=rnum,column=cnum,value = title)
    dataCell.font = titleFont
    for x in info:
        test = '' 
        dataCell = work.cell(row=rnum+1,column=cnum,value= x)
        dataCell = work.cell(row=rnum+1,column=cnum+1,value= (info[x])[0][1])
        dataCell = work.cell(row=rnum+1,column=cnum+2,value= len(info[x]))
        if len(info[x])>1:
            test = info[x][0][0]
            for y in range(1,len(info[x])):
                test+= ","+info[x][y][0]
            dataCell=work.cell(row=rnum+1,column = cnum+3,value=test)
        else:
            dataCell=work.cell(row=rnum+1,column = cnum+3,value=(info[x])[0][0])
            dataCell.font = cellFont
        rnum+=1
    dataCell = work.cell(row = rnum+1,column = 3,value = len(info)+1)

if exists("comp.txt") and exists("bom.xlsx"):
    createFiles("comp.txt")
else:
    print("No Comp.txt and BOM.xlsx found")
    input()