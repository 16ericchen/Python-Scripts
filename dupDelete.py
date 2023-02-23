import os
from openpyxl import load_workbook,Workbook
# os.chdir(os.path.dirname(__file__))
def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    dictInfo = []
    for row in sheet.iter_rows(min_row=1, min_col=1,):
        input = []
        for x in range(0,6):
            input.append(row[x].value)
        for x in range(6,sheet.max_column,2):
            if row[x].value and row[x+1].value:
                if row[x+1].value not in input:
                    input.append(row[x].value)
                    input.append(row[x+1].value)
        dictInfo.append(input)
    createXL(dictInfo)
def createXL(data):
    wx = Workbook()
    log = wx.active
    for i,x in enumerate(data):
        for y in range(len(x)):
            log.cell(row=i+1,column = y+1,value=x[y])
    wx.save(filename='BOM.xlsx')
flag = False
bomList  = []
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        flag = True
        parser(x)
if flag == False:
    print("No Excel File Found")
    input()