import os,itertools
from openpyxl import load_workbook,Workbook

os.chdir(os.path.dirname(__file__))
def expandArray(testCase):
    splitCase = testCase.split("-")
    if (splitCase[0][0:2]).isalpha() or splitCase[0][1] == "0":
        start = int(splitCase[0][2:])+1
        end = int(splitCase[1][2:])
        for x in range(start,end):
            splitCase.append(splitCase[0][0]+splitCase[0][1]+str(x))
    else:
        for x in range(int(splitCase[0][1:])+1,int(splitCase[1][1:])):
            splitCase.append(splitCase[0][0]+str(x))
    return splitCase
def addOne(num):
    carry = 0
    output = ''
    for x in reversed(range(len(num))):
        if num[x].isdigit():
            checkNum = int(num[x])+carry
            if x == len(num)-1:
                checkNum = int(num[x])+carry+1
            if checkNum < 10:
                output = str(checkNum) + output
                carry = 0
            else:
                output = str(checkNum-10) + output
                carry = 1
        else:
            if carry == 1:
                output = num[x] + '1' + output
                carry = 0
            else:
                output = num[x]+output
    return output

def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    test = []
    if sheet == None:
        print('Excel Format Cannot Be Read')
        print('Change File Type From Strict Open XML Spreadsheet To Regular XLSX Format')
        print('Then Restart the Program')
        input()
    dictInfo = {}
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=4):
        #0=part number,1=description,2=qty,3=ref des,4=level,5 = item Cat
        test.clear()
        if row[0].value == None:
            continue
        if row[3].value and 'PCB' not in row[3].value:
            refTotal = row[3].value.splitlines()
            b=[line.split(',') for line in refTotal]
            test = list(itertools.chain.from_iterable(b))
            for x in test:
                print(x)
                if "-" in x:
                    for y in expandArray(x):
                        dictInfo[y] = [row[0].value,row[1].value]   
                else:
                    dictInfo[x] = [row[0].value,row[1].value]
    dictInfo.pop(' ', None)
    dictInfo.pop('', None)
    createXL(dictInfo)
    
def createXL(data):
    wx = Workbook()
    log = wx.active
    rnum = 1
    test = list(data.keys())
    test.sort()
    for x in test:
        log.cell(row=rnum,column=1,value= x)
        log.cell(row=rnum,column=2,value= data[x][0])
        log.cell(row=rnum,column=3,value= data[x][1])
        rnum+=1
    wx.save(filename='bom.xlsx')

    # while cell
flag = False
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        flag = True
        parser(x)
    if x.endswith(('.xls')):
        print('Change File Type from XLS to XLSX Format')
        print('Then Restart the Program')
        input()
if flag == False:
    print("No Excel File Found")
    input()