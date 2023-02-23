import os,itertools
from openpyxl import load_workbook,Workbook
from collections import OrderedDict

os.chdir(os.path.dirname(__file__))

i = 2
listofBOMS = []
component = {}
masterList = {}
cur = ''
newKey = ''

def expandArray(testCase):
    splitCase = testCase.split("-")
    for i,j in enumerate(splitCase[0]):
        if not j.isalpha() and j != '0':
            pointer = i
            break
    start = splitCase[0]
    end = splitCase[1]
    small = int(start[pointer:])
    big = int(end[pointer:])
    for x in range(small+1,big):
        splitCase.append(start[:pointer]+str(x))
    return splitCase

def createEngBOM(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    test = []
    if sheet == None:
        print('Excel Format Cannot Be Read')
        print('Change File Type From Strict Open XML Spreadsheet To Regular XLSX Format')
        print('Then Restart the Program')
        input()
    dictInfo = {}
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=9):
        #1=part number,2=description,3=qty,6=ref des,0=level,8=item Cat
        test.clear()
        if row[0].value == None:
            continue
        if row[6].value and 'PCB' not in row[6].value:
            refTotal = row[6].value.splitlines()
            b=[line.split(',') for line in refTotal]
            test = list(itertools.chain.from_iterable(b))
            for x in test:
                if "-" in x:
                    for y in expandArray(x):
                        dictInfo[y] = [row[1].value,row[2].value]   
                else:
                    dictInfo[x] = [row[1].value,row[2].value]
    dictInfo.pop(' ', None)
    dictInfo.pop('', None)
    return dictInfo
def parser(data,count):
    for x in data:
        if x in component.keys():
            if len(component[x]) == count:
                component[x] = component[x]+[data[x][0]]
            else:
                fill = ['No Stuff']*(count-len(component[x]))
                component[x] = component[x]+fill+[data[x][0]]
        else:
            if count>0:
                fill = ['No Stuff']*count
                component[x] = fill + [data[x][0]]
            else:
                component[x] = [data[x][0]]

def locationDifference():
    for x in os.listdir(os.getcwd()):
        if x.endswith(('.xlsx')) and x !='BOMComparison.xlsx':
            listofBOMS.append(x)
    wb = Workbook()
    log = wb.active
    log.title = 'BOMComparison'
    for x in range(len(listofBOMS)):
        log.cell(row=1,column=x+4,value=listofBOMS[x][:-5])
        test=createEngBOM(listofBOMS[x])
        parser(test,x)
    final= OrderedDict(sorted(component.items()))
    for x,y in final.items():
        if len(y)<len(listofBOMS):
            fill = ['No Stuff']*(len(listofBOMS)-len(y))
            y = y + fill
        if len(set(y))>1:
            log.cell(row=i,column=3,value=x)
            for z in range(len(y)):
                log.cell(row=i,column=z+4,value=y[z])
            i +=1
    wb.save(filename='BOMComparison.xlsx')