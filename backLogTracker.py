
from datetime import datetime
from openpyxl import load_workbook,Workbook
import os
from os import path
from os.path import exists
from openpyxl.utils import get_column_letter
poInfo = {}
headers = ["Material",'Customer Part','QTY','Estimated Ship Date','Revised Estimated Ship Date','Ship Date','Tracking Ref']
# os.chdir(path.dirname(__file__))
def main(inputFile,xlFile):
    f = open(inputFile,"r",encoding='utf-8')
    readCompFile(f)
    readBomFile(xlFile)
    if exists("PO_Info.xlsx"):
        exportCAD('PO Info-2.xlsx')
    else:
        exportCAD('PO_Info.xlsx')
    input()
def exportCAD(fileName):
    wb = Workbook()
    for name in poInfo.keys():
        wb.create_sheet(name,0)
        ws =wb[name]
        rnum = 2
        for m in range(len(headers)):
            ws.cell(row=rnum-1,column = m+1, value = headers[m])
            ws.column_dimensions[get_column_letter(m+1)].width = 15
        for y in poInfo[name]:
            for x in range(len(y)):
                if y[x] == None:
                    ws.cell(row=rnum,column=x+1,value= ' ')
                else:
                    if type(y[x]) == datetime:
                        ws.cell(row=rnum,column=x+1,value= y[x].date())
                    else:
                        ws.cell(row=rnum,column=x+1,value= y[x])
            rnum+=1
    wb.save(filename=fileName)

def readCompFile(inputArray):
    for line in inputArray:
        poInfo[line.strip()] = []
    inputArray.close()
def readBomFile(exlFile):
    wb = load_workbook(filename=exlFile)
    sheet = wb.active
    for row in sheet.iter_rows(min_row = 2,max_row = None,min_col=1,max_col=None):
        if row[11].value in poInfo:
            poInfo[row[11].value].append([row[9].value,row[10].value,row[17].value,row[19].value,row[20].value,row[21].value,row[31].value])

for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')) and 'PO_Info' not in x:
        xl = x
    if x.endswith(('.xls')):
        print('Change File Type from XLS to XLSX Format')
        print('Then Restart the Program')
        input()
    if x.endswith(('.txt')):
        tx = x
if tx and xl:
    main(tx,xl) 
else:
    print("Files Not Found")