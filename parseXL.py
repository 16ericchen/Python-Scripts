import os
from openpyxl import load_workbook


apnDict = {}

def subFolders(path,names):
    for x in names:
        for y in os.listdir(path+'\\'+x):
            if y.endswith(('.xlsx')) and "~" not in y:
                parseXL(path+'\\'+x+'\\'+y,x) 


def parseXL(file,name):
    wb = load_workbook(filename=file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=3, min_col=1):
        if row[0].value not in apnDict.keys():
            apnDict[row[0].value] = [name]
        else:
            apnDict[row[0].value].append(name)




path = r"C:\Users\EricChen\Desktop\Python Scripts\In Progress\buyerDatabase\APN Database based on project"
folders = next(os.walk(path))[1]

for x in folders:
    sub = next(os.walk(path+"\\"+x))[1]
    if 'OLD' in sub:
        sub.remove('OLD')
    if '.DAV' in sub:
        sub.remove('.DAV')
    if sub:
        subFolders(path+"\\"+x,sub)
    else:
        for y in os.listdir(path+'\\'+x):
            if y.endswith(('.xlsx')) and "~" not in y:
                parseXL(path+'\\'+x+'\\'+y,x)