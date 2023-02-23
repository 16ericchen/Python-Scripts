from openpyxl import load_workbook,Workbook
import os
from collections import OrderedDict
os.chdir(os.path.dirname(__file__)) 


#This one is deleting some parts when shortening


def parser(file,count):
    wb = load_workbook(filename=file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, min_col=1):
        if row:
            if row[0].value in component.keys():
                if len(component[row[0].value]) == count:
                    component[row[0].value] = component[row[0].value]+[row[1].value]
                else:
                    fill = ['No Stuff']*(count-len(component[row[0].value]))
                    component[row[0].value] = component[row[0].value]+fill+[row[1].value]
            else:
                if count>0:
                    fill = ['No Stuff']*count
                    component[row[0].value] = fill + [row[1].value]
                else:
                    component[row[0].value] = [row[1].value]
        else:
            break
i = 2
listofBOMS = []
component = {}
masterList = {}
cur = ''
newKey = ''
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')) and x !='BOMComparison.xlsx':
        listofBOMS.append(x)
wb = Workbook()
log = wb.active
log.title = 'BOMComparison'
for x in range(len(listofBOMS)):
    log.cell(row=1,column=x+4,value=listofBOMS[x][-10:-5])
    parser(listofBOMS[x],x)

for x in component.keys():
    if component[x] == cur:
        newKey = newKey + ','+ x
    else:
        if cur!='' and newKey!='':
            masterList[newKey] = cur
        cur = component[x]
        newKey = x
final= OrderedDict(sorted(masterList.items()))
for x,y in final.items():
    if len(y)<len(listofBOMS):
        fill = ['No Stuff']*(len(listofBOMS)-len(y))
        y = y + fill
    if len(set(y))>1:
        log.cell(row=i,column=3,value=x)
        for z in range(len(y)):
            log.cell(row=i,column=z+4,value=y[z])
        i +=1

wb.save(filename='BOMComparison.xlsx')

