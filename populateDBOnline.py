import sqlite3
from sqlite3 import Error
import os
from os import path
from openpyxl import load_workbook
import warnings
warnings.simplefilter("ignore")
apnDict = {}

# os.chdir(path.dirname(__file__))
def enterData():
    database = r"N:\Python\Project Database\APN.db"
    conn = create_connection(database)
    for x in apnDict:
        projects = ','.join(apnDict[x])
        task_1 = x,projects
        create_projects(conn,task_1)
    print("Done")

def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except Error as e:
        print(e)
    return conn

def create_table(conn, create_table_sql):
    try:
        c = conn.cursor()
        c.execute(create_table_sql)
    except Error as e:
        print(e)

def create_projects(conn, projects):
    sql = ''' INSERT or IGNORE INTO projects(APN,Project_Name)
              VALUES(?,?) '''
    cur = conn.cursor()
    cur.execute(sql, projects)
    conn.commit()
    return cur.lastrowid

def createTable():
    database = r"N:\Python\Project Database\APN.db"
    conn = create_connection(database)
    sql_create_projects_table = """ CREATE TABLE IF NOT EXISTS projects (
                                            APN text PRIMARY KEY,
                                            Project_Name text
                                        ); """
    create_table(conn, sql_create_projects_table)

def subFolders(path,names):
    for x in names:
        for y in os.listdir(path+'\\'+x):
            if y.endswith(('.xlsx')) and "~" not in y:
                parseXL(path+'\\'+x+'\\'+y,x) 
                
def parseXL(file,name):
    wb = load_workbook(filename=file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=3, min_col=1):
        if row[0].value not in apnDict.keys():
            apnDict[row[0].value] = [name]
        else:
            apnDict[row[0].value].append(name)

def getData():
    path = r"N:\WHS\Project Inventory"
    folders = next(os.walk(path))[1]
    for x in folders:
        sub = next(os.walk(path+"\\"+x))[1]
        if 'OLD' in sub:
            sub.remove('OLD')
        if '.DAV' in sub:
            sub.remove('.DAV')
        if sub:
            subFolders(path+"\\"+x,sub)
        else:
            for y in os.listdir(path+'\\'+x):
                if y.endswith(('.xlsx')) and "~" not in y:
                    parseXL(path+'\\'+x+'\\'+y,x)
def main():
    database = r"N:\Python\Project Database\APN.db"
    getData()
    conn = create_connection(database)
    createTable()
    if conn is not None:
        enterData()
    else:
        print("Error! cannot create the database connection.")

if __name__ == '__main__':
    main()

