import os
import shutil
from openpyxl import load_workbook
def aoi():
    projectListPath = r"C:\Users\EricChen\Desktop\SPI Need To Do\SPI Program List.xlsx"
    projectInfo = fullName =  ''
    wb = load_workbook(projectListPath)
    print("Input Board Number:")
    boardNumber = input()
    ws = wb.active
    for cell in ws['A']:
        if boardNumber in cell.value:
            projectInfo = cell.value
    if projectInfo == '':
        print("Board Number Not Found")
        aoi()
    projectInfo = projectInfo.replace(',','')
    folderPath = r"C:\Users\EricChen\Desktop\AOI"+'\\'+projectInfo
    if not os.path.exists(folderPath):
        os.makedirs(folderPath)
    projectLetter = projectInfo[0]
    test = projectInfo.split(' ')
    projectName = ' '.join(test[0:-2])
    projectNumber = test[0]
    projectDate = test[-1]
    projectFolderPath = "K:\Projects"+"\\"+projectLetter+"\\"+projectNumber
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        aoi()
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if x.lower() in projectName.lower():
            fullName =projectFolderPath+ "\\"+x+"\\"+projectName+" "+projectDate
    if fullName == '':
        print("File Path Not Found")
        aoi()
    projectFolderPath = fullName
    print(fullName)
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        aoi()
    compPath = projectFolderPath + '\\'+'Gerber'+'\\'+'comp.txt'
    if not os.path.exists(compPath):
        print("Comp File Not Found")
        aoi()
    if os.path.exists(compPath):
        shutil.copy(compPath,folderPath)
        print("Comp File was Moved")
    bomPath = projectFolderPath + '\\'+'BOM_CAD'
    for x in os.listdir(bomPath):
        if 'bom' in x.lower():
            shutil.copy(bomPath+"\\"+x,folderPath)
            print("BOM File was Moved")
    buildMatrixPath = projectFolderPath +'\\'+'Information'
    for x in os.listdir(buildMatrixPath):
        if x.endswith(('.xlsx','xls')):
            shutil.copy(buildMatrixPath+"\\"+x,folderPath)
        print("Build Matrix File was Moved")
    aoi()
aoi()
