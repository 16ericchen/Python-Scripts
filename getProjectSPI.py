import os
import shutil
from openpyxl import load_workbook
def spi():
    projectListPath = r"C:\Users\EricChen\Desktop\SPI Need To Do\SPI Program List.xlsx"
    projectInfo = ''
    wb = load_workbook(projectListPath)
    print("Input Board Number:")
    boardNumber = input()
    ws = wb.active
    for cell in ws['A']:
        if boardNumber in cell.value:
            projectInfo = cell.value
    if projectInfo == '':
        print("Board Number Not Found")
        spi()
    projectInfo = projectInfo.replace(',','')
    spiNTDPath = r"C:\Users\EricChen\Desktop\SPI Need To Do"+'\\'+projectInfo
    spiTransferPath = r"C:\Users\EricChen\Desktop\SPI Transfer"+'\\'+projectInfo
    if not os.path.exists(spiNTDPath):
        os.makedirs(spiNTDPath)
    projectLetter = projectInfo[0]
    projectInfo = projectInfo.replace(',','')
    test = projectInfo.split(' ')
    projectName = ' '.join(test[0:-2])
    projectNumber = test[0]
    projectDate = test[-1]
    projectFolderPath = "K:\Projects"+"\\"+projectLetter+"\\"+projectNumber
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        spi()
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if x.lower() in projectName.lower():
            fullName =projectFolderPath+ "\\"+x+"\\"+projectName+" "+projectDate
    projectFolderPath = fullName
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        spi()
    compPath = projectFolderPath + '\\'+'Gerber'+'\\'+'comp.txt'
    if not os.path.exists(compPath):
        print("Comp File Not Found")
        spi()
    if os.path.exists(compPath):
        shutil.copy(compPath,spiNTDPath)
        print("COMP File was Moved")
    designPath = projectFolderPath + '\\'+'Download'+'\\'+boardNumber+'.tgz'
    if not os.path.exists(designPath):
        print("TGZ File Not Found")
        spi()
    if os.path.exists(designPath):
        shutil.copy(designPath,spiNTDPath)
        print("TGZ File was Moved")
    if not os.path.exists(spiTransferPath+"\TOP"):
            os.makedirs(spiTransferPath+"\TOP")
    if not os.path.exists(spiTransferPath+"\BOT"):
            os.makedirs(spiTransferPath+"\BOT")
    spi()
spi()
