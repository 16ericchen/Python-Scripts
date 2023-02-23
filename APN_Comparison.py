from openpyxl import load_workbook,Workbook
import os
from os import path

listofBOMS = []
status = {}
matcheS = []
info = {}
os.chdir(path.dirname(__file__))

def enterData():
    for x in os.listdir(os.getcwd()):
        if x.endswith(('.xlsx')) and x !='BOMComparison.xlsx':
            listofBOMS.append(x)
    wb = Workbook()
    log = wb.active
    log.title = 'BOMComparison'
    for x in listofBOMS:
        parser(x)
    subsets = [[]]
    for n in list(status.keys()):
        subsets += [s + [n] for s in subsets]
    new_list = [com for com in subsets if len(com)>1]
    new_list = reversed(new_list)
    for x in new_list:
        intersectionExport(x)
    export()

def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    start_row = 0
    status[file[:-5]] = []
    for row in sheet.iter_rows(min_row=1, min_col=1,max_col=1):
        if row[0].value is not None:
            if row[0].value.lower() == 'level':
                start_row = row[0].row
                break
    for row in sheet.iter_rows(min_row = start_row+1,min_col = 2,max_col=2):
        if row[0].value:
            status[file[:-5]].append(row[0].value)

def intersectionExport(x):
    same = status[x[0]]
    maxLen = 0
    for dataSets in x:
        maxLen = max(maxLen,len(status[dataSets]))
        same = set(same).intersection(status[dataSets])
    if len(same) in info.keys():
        if info[len(same)][0][0] == same:
            info[len(same)][1].append(x)
    else:
        info[len(same)] = [[same,maxLen],[x]]

    

def export():
    wb = Workbook()
    log = wb.active
    log.title = 'BOMComparison'
    nums = sorted(info.keys(),reverse=True)
    row = 3
    col = 3
    row2 = 3
    for x in nums:
        row = 3
        for combo in info[x][1]:
            log.cell(row=row,column=col,value=','.join(combo))
            row +=1
        log.cell(row=row+1,column=col,value=str(len(info[x][0][0]))+'/'+str(info[x][0][1]))
        col+=2
    col = 3
    for x in nums:
        row2 = 3
        for y in info[x][0][0]:
            log.cell(row=row+row2,column=col,value=y)
            row2 +=1
        col+=2
    wb.save(filename='APNComparison.xlsx')  

enterData()