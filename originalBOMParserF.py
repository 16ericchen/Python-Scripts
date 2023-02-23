# Delete row if level is not 1
# Delete row if Item Cat is not L
# need part number,description,and Ref Des
# output ref des,part number, description in this order
import os
from openpyxl.utils import column_index_from_string
from openpyxl import load_workbook,Workbook
def parser(file):
    print('parsed'+': '+file)
    wb = load_workbook(filename=file)
    sheet = wb.active
    if sheet == None:
        print('Excel Format Cannot Be Read')
        print('Change File Type From Strict Open XML Spreadsheet To Regular xlsx Format')
        print('Then Restart the Program')
        input()
    for y in wb.sheetnames:
        getCord(wb[y],y)

def getCord(sheet,sheetName):
    level = ''
    ar = {}
    for row in sheet.iter_cols(min_row=1, min_col=1, max_row=100, max_col=50):
        for cell in row:
            if "Level" == str(cell.value):
                level = cell.coordinate
                break
    for row in sheet.iter_cols(min_col=1,max_col=None,min_row=int(level[1:]),max_row=int(level[1:])):
        if len(ar) == 4:
            break
        for cell in row:
            if cell.value == None and len(ar)<4:
                continue
            if 'Part Number' in cell.value and 'Part Number' not in ar or 'Item Number' in cell.value:
                ar["Part Number"]=cell.column
            if 'Description' in cell.value and 'Description' not in ar:
                ar["Description"] = cell.column
            if 'Ref Des' in cell.value and 'Ref Des' not in ar:
                ar["Ref Des"] = cell.column
            if 'Item Cat' in cell.value and 'Item Cat' not in ar:
                ar["Item Cat"] = cell.column
            if len(ar) == 4:
                break
            else:
                continue
    dictInfo = {}
    i = int(level[1:])+1
    col = column_index_from_string(level[0])
    # use ref des as key for dictionary and split ref des so loop through it
    for i in range(1,sheet.max_row+1):
        if sheet.cell(row=i,column=1).value != None:
            if '1' in str(sheet.cell(row=i,column=col).value) and sheet.cell(row=i,column=ar["Item Cat"]).value=='L' and 'PCB' not in sheet.cell(row=i,column = ar['Ref Des']).value:
                refTotal = sheet.cell(row=i,column=ar["Ref Des"]).value.split(',')
                for x in refTotal:
                    dictInfo[x] = [sheet.cell(row=i,column=ar["Part Number"]).value,sheet.cell(row=i,column=ar["Description"]).value]
    createXL(dictInfo,sheetName)

def createXL(data,name):
    wx = Workbook()
    log = wx.active
    log.title = name
    rnum = 1
    for x,y in data.items():
        log.cell(row=rnum,column=1,value= x)
        log.cell(row=rnum,column=2,value= y[0])
        log.cell(row=rnum,column=3,value= y[1])
        rnum+=1
    wx.save(filename='bom-'+name+'.xlsx')

flag = False
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        flag = True
        parser(x)
    if x.endswith(('.xls')):
        print('Change File Type from XLS to XLSX Format')
        print('Then Restart the Program')
        input()
if flag == False:
    print("No Excel File Found")
    input()