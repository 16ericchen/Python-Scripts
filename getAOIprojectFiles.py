import os
import shutil
from openpyxl import load_workbook
def aoi():
    desktop = os.path.join(os.environ['USERPROFILE'],'Desktop','AOI')  
    projectListPath = r"K:\Testing\SPI\SPI Program List.xlsx"
    simpleName = projectInfo = fullName =  ''
    wb = load_workbook(projectListPath)
    print("Input Board Number:")
    boardNumber = input()
    ws = wb.active
    for cell in ws['A']:
        if boardNumber in cell.value:
            projectInfo = cell.value
    if projectInfo == '':
        print("Board Number Not Found")
        aoi()
    projectInfo = projectInfo.replace(',','')
    if not os.path.exists(desktop):
        os.makedirs(desktop)
    folderPath = os.path.join(desktop,projectInfo)
    if not os.path.exists(folderPath):
        os.makedirs(folderPath)
    projectLetter = projectInfo[0]
    test = projectInfo.split(' ')
    test = [i for i in test if i]
    projectName = ' '.join(test[0:-2])
    if 'board' in projectInfo:
        simpleName = projectInfo.split('board')[0]
    if 'board' not in projectInfo and simpleName == '':
        simpleName = projectInfo.split('Proto')[0]
    projectNumber = test[0]
    projectDate = test[-1]
    projectFolderPath = "K:\Projects"+"\\"+projectLetter+"\\"+projectNumber
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        aoi()
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if x.replace(' ','').lower() in projectName.replace(' ','').lower() or simpleName.replace(' ','').lower() in x.replace(' ','').lower():
            fullName =projectFolderPath+ "\\"+x
            if simpleName.replace(' ','').lower() == x.replace(' ','').lower():
                break
    projectFolderPath = fullName
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        aoi()
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if (x[:-11] in projectName or projectDate==folders[0][-10:]) and x[:-11] != '':
            if x[:-11] == projectName:
                projectFolderPath = projectFolderPath +"\\"+x
                break
    print(projectFolderPath)
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        aoi()
    compPath = projectFolderPath + '\\'+'Gerber'+'\\'+'comp.txt'
    if not os.path.exists(compPath):
        print("Comp File Not Found")
        aoi()
    if os.path.exists(compPath):
        shutil.copy(compPath,folderPath)
        print(compPath)
        print("Comp File was Moved")
    designPath = projectFolderPath + '\\'+'Download'+'\\'+boardNumber+'.tgz'
    if not os.path.exists(designPath):
        print("TGZ File Not Found")
        aoi()
    if os.path.exists(designPath):
        shutil.copy(designPath,folderPath)
        print("TGZ File was Moved")
        print(designPath)
    bomPath = projectFolderPath + '\\'+'BOM_CAD'
    for x in os.listdir(bomPath):
        if 'bom' in x.lower():
            shutil.copy(bomPath+"\\"+x,folderPath)
            print("BOM File was Moved")
    buildMatrixPath = projectFolderPath +'\\'+'Information'
    for x in os.listdir(buildMatrixPath):
        if x.endswith(('.xlsx','xls')):
            shutil.copy(buildMatrixPath+"\\"+x,folderPath)
        print("Build Matrix File was Moved")
    
    aoi()
aoi()
