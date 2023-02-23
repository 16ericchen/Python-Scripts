

# then read through each column and get the row number for cells with a value
# seperate the values between no stuff and stuff and null
# when iterating through config column if there is a value then combine applepncolumn variable with current row number to get value repeat with ref_des 
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

apn = ["APN","Apple PN","Part#","Part #","P/N"]
def buildMatrix():
    global str_list
    nostuffList,stuffList = [],[]
    applePN=configL=''
    projectListPath = r"C:\Users\EricChen\Desktop\Python Scripts\Build Matrix Monarch Rally Dev1 SMA9 05-04-22 REV2.xlsx"
    wb = load_workbook(projectListPath)
    sheet = wb.active
    for row in sheet.iter_cols(min_row=1, min_col=1, max_row=100, max_col=50):
        for cell in row:
            if cell.value in apn and applePN == '':
                applePN = cell.coordinate
            # difference between blue and yellow excel grabber is the cell.value condition check being Configs for blue and Config for Yellow
            if cell.value == "Configs" and configL == '':
                configL = cell.coordinate
    print("Input Config:")
    configInput = input()
    configCol = column_index_from_string(coordinate_from_string(configL)[0])+1
    for row in sheet.iter_cols(min_row=int(configL[1]),min_col=configCol,max_row=int(configL[1]),max_col=100):
        if row[0].value==configInput:
            configCol = row[0].column
            break
    applePNCol = column_index_from_string(coordinate_from_string(applePN)[0])
    for row in sheet.iter_rows(min_row = int(applePN[1:]),min_col = configCol,max_row = 100,max_col = configCol):
        if not row[0].value:
            continue
        elif "no stuff" in (row[0].value).lower():
            nostuffList.append(sheet.cell(row[0].row,applePNCol).value.replace(' ',''))
        else:
            stuffList.append(row[0].value)
buildMatrix()


