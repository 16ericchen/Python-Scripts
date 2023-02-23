import os
from openpyxl import load_workbook,Workbook
os.chdir(os.path.dirname(__file__))
from decimal import Decimal
from datetime import date
titles = ["Plant","Storage location","Material","Description","Quantity","Net Price","Price Unit","Material group","Free Item","Item category","Account assignment category","Cost Center","G/L account","WBS element","Mfr","Mfr part no.","Item text"]
distributor0101 = ["ARROW","DIGIKEY","MOUSER","SAMTEC","TI","AKM","AT&S","CCTC","COMPEQ","NANYA","TRIPOD","UMC","UNITECH","ZDT"]
distributorDict = {}

def createPO(manufacturer,info):
    today = date.today()
    fileName = manufacturer[0]+' '+info[0][13]+info[0][16]+" "+today.strftime("%m%d%y")+".xlsx"
    wb = Workbook()
    log = wb.active
    log.title = "PO"
    i = 0
    for i,x in enumerate(titles):
        log.cell(row=1,column=i+1,value=x)
    for i,x in enumerate(info):
        for y in range(len(x)):
            log.cell(row=i+2,column=y+1,value=x[y])
    wb.save(filename=fileName)
      
def findNetPrice(value):
    w = Decimal(value)
    test = 0
    if '.' not in str(w): 
        return [w,1]
    test = str(w).index('.')
    if len(str(w))-test>3:
        magnitude = 10**(len(str(w))-test-3)
        w*=magnitude
        if magnitude >10000:
            return [w,'Exceed Max']
        return[w,magnitude]
    else:
        return [w,1]

def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    #Find Start Row and the column for each necessary value
    for row in sheet.iter_rows(min_row=1,min_col=1,max_col=8):
        if row[7].value != None:
            curText = row[7].value
        plant = '0126'
        storage = ' '
        material = 'MG20'
        account = 'K'
        LGAccount = '139901'
        itemText = curText
        WBS = ''
        if row[0].value:
            price = findNetPrice(str(row[4].value))
            if (row[3].value.replace('-','').split(' ')[0].upper()) in distributor0101 and (row[3].value.upper() !='TI STORE'):
                print(row[3].value)
                plant = 'ZSJ1'
                storage = '1000'
                material = ''
                account = 'Q'
                LGAccount = ''
                WBS = curText
                itemText = ''
            if itemText:
                if (row[3].value.split(' ')[0],itemText) not in distributorDict.keys():
                    distributorDict[(row[3].value.split(' ')[0],itemText)] = [[plant,storage,row[0].value,row[1].value,row[2].value,price[0],price[1],material,'','',account,'BV1057',LGAccount,WBS,row[5].value,row[6].value,itemText]]
                else:
                    distributorDict[(row[3].value.split(' ')[0],itemText)].append([plant,storage,row[0].value,row[1].value,row[2].value,price[0],price[1],material,'','',account,'BV1057',LGAccount,WBS,row[5].value,row[6].value,itemText])
            else:
                if (row[3].value.split(' ')[0],WBS) not in distributorDict.keys():
                    distributorDict[(row[3].value.split(' ')[0],WBS)] = [[plant,storage,row[0].value,row[1].value,row[2].value,price[0],price[1],material,'','',account,'BV1057',LGAccount,WBS,row[5].value,row[6].value,itemText]]
                else:
                    distributorDict[(row[3].value.split(' ')[0],WBS)].append([plant,storage,row[0].value,row[1].value,row[2].value,price[0],price[1],material,'','',account,'BV1057',LGAccount,WBS,row[5].value,row[6].value,itemText])
    for x in distributorDict:
        createPO(x,distributorDict[x])

flag = False
bomList  = []
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        flag = True
        parser(x)
        break
if flag == False:
    print("No Excel File Found")
    input()















