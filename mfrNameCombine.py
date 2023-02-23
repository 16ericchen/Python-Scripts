import os
from openpyxl import load_workbook,Workbook

os.chdir(os.path.dirname(__file__))
def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    dictInfo = {}
    cur = ' '
    curMax = 0
    currentMNF = ''
    for row in sheet.iter_rows(min_row=3, min_col=1, max_col=8):
        if row[6].value:
            currentMNF = row[6].value
        if row[1].value:
            if row[5].value:
                dictInfo[row[1].value] = [row[5].value,row[3].value,row[2].value,row[4].value,row[0].value,row[6].value,row[7].value]
                cur = row[1].value
                currentMNF = row[6].value
            else:
                dictInfo[row[1].value] = [row[5].value,row[3].value,row[2].value,row[4].value,row[0].value,' ',' ']
        else:
            if row[6].value:
                dictInfo[cur]+=[row[6].value,row[7].value]
                curMax = max(len(dictInfo[cur]),curMax)
                currentMNF = row[6].value
            elif row[7].value:
                print('here')
                print(row[7].value)
                dictInfo[cur]+=[currentMNF,row[7].value]
                curMax = max(len(dictInfo[cur]),curMax)
            else:
                break
    createXL(dictInfo,curMax)
def createXL(data,maxL):
    wx = Workbook()
    log = wx.active
    rnum = 2
    col =1
    log.cell(row=1,column=1,value = 'Ref Des')
    log.cell(row=1,column=2,value = 'BOM.Item Cat')
    log.cell(row=1,column=3,value = 'APN')
    log.cell(row=1,column=4,value = 'Description')
    log.cell(row=1,column=5,value = 'Qty')
    log.cell(row=1,column=6,value = 'Level')
    
    for l in range(7,maxL+1):
        if l%2==1:
            log.cell(row=1,column=l,value = 'Manufacturer Part Number') 
        else:
            log.cell(row=1,column=l,value = 'Manufacturer Name') 
    for x,y in data.items():
        log.cell(row=rnum,column=3,value= x)
        if len(y)>1:
            for i in range(len(y)):
                if i != 2:
                    log.cell(row=rnum,column=col,value= y[i])
                    col+=1
                else:
                    col+=1
                    log.cell(row=rnum,column=col,value= y[i])
                    col+=1
                    continue
        else:
            log.cell(row=rnum,column=2,value= y[0])
        rnum+=1
        col = 1
    wx.save(filename='BOM.xlsx')
flag = False
bomList  = []
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        flag = True
        parser(x)

if flag == False:
    print("No Excel File Found")
    input()