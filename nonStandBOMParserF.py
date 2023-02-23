import os,itertools
from openpyxl import load_workbook,Workbook

os.chdir(os.path.dirname(__file__))
def expandArray(testCase):
    splitCase = testCase.split("-")
    for i,j in enumerate(splitCase[0]):
        if not j.isalpha() and j != '0':
            pointer = i
            break
    start = splitCase[0]
    end = splitCase[1]
    small = int(start[pointer:])
    big = int(end[pointer:])
    for x in range(small+1,big):
        splitCase.append(start[:pointer]+str(x))
    return splitCase

def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    test = []
    if sheet == None:
        print('Excel Format Cannot Be Read')
        print('Change File Type From Strict Open XML Spreadsheet To Regular XLSX Format')
        print('Then Restart the Program')
        input()
    dictInfo = {}
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=4):
        #0=part number,1=description,2=qty,3=ref des,4=level,5 = item Cat
        test.clear()
        if row[0].value == None:
            continue
        if type((row[3].value)) == int:
            print("Check Row",row[3].row,"the location is wrong")
            print("Please change and restart program")
            input()
        if row[3].value and 'PCB' not in row[3].value:
            refTotal = row[3].value.splitlines()
            b=[line.split(',') for line in refTotal]
            test = list(itertools.chain.from_iterable(b))
            for x in test:
                if "-" in x:
                    for y in expandArray(x):
                        dictInfo[y] = [row[0].value,row[1].value]   
                else:
                    dictInfo[x] = [row[0].value,row[1].value]
    dictInfo.pop(' ', None)
    dictInfo.pop('', None)
    createXL(dictInfo)
    
def createXL(data):
    wx = Workbook()
    log = wx.active
    rnum = 1
    test = list(data.keys())
    test.sort()
    for x in test:
        log.cell(row=rnum,column=1,value= x)
        log.cell(row=rnum,column=2,value= data[x][0])
        log.cell(row=rnum,column=3,value= data[x][1])
        rnum+=1
    wx.save(filename='bom.xlsx')

    # while cell
flag = False
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        flag = True
        parser(x)
    if x.endswith(('.xls')):
        print('Change File Type from XLS to XLSX Format')
        print('Then Restart the Program')
        input()
if flag == False:
    print("No Excel File Found")
    input()