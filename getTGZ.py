import os
import shutil
from openpyxl import load_workbook
def spi():
    desktop = os.path.join(os.environ['USERPROFILE'],'Desktop','TGZ') 
    projectListPath = r"K:\Testing\SPI\SPI Program List.xlsx"
    projectInfo = ''
    wb = load_workbook(projectListPath)
    print("Input Board Number:")
    boardNumber = input()
    ws = wb.active
    for cell in ws['A']:
        if boardNumber in cell.value:
            projectInfo = cell.value
    if projectInfo == '':
        print("Board Number Not Found")
        spi()
    projectInfo = projectInfo.replace(',','')
    if not os.path.exists(desktop):
        os.makedirs(desktop)
    spiNTDPath = os.path.join(desktop,projectInfo)
    if not os.path.exists(spiNTDPath):
        os.makedirs(spiNTDPath)
    projectLetter = projectInfo[0]
    projectInfo = projectInfo.replace(',','')
    test = projectInfo.split(' ')
    projectName = ' '.join(test[0:-2])
    projectNumber = test[0]
    projectDate = test[-1]
    projectFolderPath = "K:\Projects"+"\\"+projectLetter+"\\"+projectNumber
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        spi()
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if x.lower() in projectName.lower():
            fullName =projectFolderPath+ "\\"+x
    projectFolderPath = fullName
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        spi()
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if projectName == x[:-11] or projectDate==folders[0][-10:]:
            projectFolderPath = projectFolderPath +"\\"+x
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        spi()
    designPath = projectFolderPath + '\\'+'Download'+'\\'+boardNumber+'.tgz'
    if not os.path.exists(designPath):
        print("TGZ File Not Found")
        spi()
    if os.path.exists(designPath):
        shutil.copy(designPath,spiNTDPath)
        print("TGZ File was Moved")
        print(designPath)
    spi()
spi()
