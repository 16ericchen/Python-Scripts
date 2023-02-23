import sqlite3,datetime
from sqlite3 import Error
from consolemenu import *
from consolemenu.items import *
from openpyxl import Workbook

totalList = [] 
time = datetime.date.today()
projectDate = time.strftime("%m-%d-%Y")
def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except Error as e:
        print(e)
    return conn

def create_table(conn, create_table_sql):
    try:
        c = conn.cursor()
        c.execute(create_table_sql)
    except Error as e:
        print(e)

def exportExcel():
    database = r"C:\Users\EricChen\Desktop\Shipping Database\testDatabase.db"
    wb = Workbook()
    dest_filename = 'daily_shipping_log'+"("+projectDate+")"+'.xlsx'
    log = wb.active
    log.title = projectDate
    conn = create_connection(database)
    cur = conn.cursor()
    cur.execute("SELECT * FROM projects WHERE date=?", (projectDate,))
    x = cur.fetchall()
    inputItems = ["Project","Assembly","Serial Number","Matrix","Note"]
    createTemplate(log,len(x),inputItems)
    for y in range(0,(len(x)*5),5):
        log.cell(column = y+1, row = 2, value = x[y//5][0])
        log.cell(column = y+2, row = 2, value = x[y//5][1])
        print(x[y//5][0])
        cur.execute("SELECT serialNumber,matrix,note FROM serialTable WHERE (project_name,assembly_name,date) = (?,?,?)",(x[y//5][0],x[y//5][1],x[y//5][2]))
        i = cur.fetchall()
        addSerial(y,i,log)
    wb.save(filename=dest_filename)

def addSerial(columnNumber,values,sheet):
    for x in range(2,len(values)+2):
        sheet.cell(column = columnNumber+3, row = x, value = values[x-2][0])
        sheet.cell(column = columnNumber+4, row = x, value = values[x-2][1])
        sheet.cell(column = columnNumber+5, row = x, value = values[x-2][2])

def createTemplate(workbook,amount,inputItems):
    i = 0 
    for x in range(1,amount*5+1):
        workbook.cell(column = x ,row = 1,value = inputItems[i])
        if i == 4:
            i = 0
        else:
            i+=1

def create_projects(conn, projects):
    sql = ''' INSERT INTO projects(name,assembly,date)
              VALUES(?,?,?) '''
    cur = conn.cursor()
    cur.execute("SELECT * FROM projects WHERE (name,assembly,date) = (?,?,?)",(projects[0],projects[1],projects[2],))
    x = cur.fetchall()
    if not x:
        cur.execute(sql, projects)
    conn.commit()
    print(cur.lastrowid)
    return cur.lastrowid

def create_serial(conn, projects):
    sql = ''' INSERT INTO serialTable(project_name,assembly_name,matrix,note,serialNumber,date)
              VALUES(?,?,?,?,?,?) '''
    cur = conn.cursor()
    cur.execute("SELECT * FROM serialTable WHERE serialNumber=?", (projects[4], ))
    x =cur.fetchone()
    print(x)
    if x:
        print("Serial Number Already Exists:")
        print("Project Date: "+x[-1]+', Project Name: '+x[0]+", Project Assembly: "+x[1]+", Serial Number: "+x[2]+", Project Matrix: "+x[3]+", Project Note: "+x[4])
    else:
        cur.execute(sql, projects)
    conn.commit()
    print("Return",cur.lastrowid)
    return cur.lastrowid

def deleteEntry(conn,project):
    cur = conn.cursor()
    cur.execute('DELETE FROM serialTable WHERE serialNumber=?', (project[4],))
    conn.commit()

def createTable():
    database = r"C:\Users\EricChen\Desktop\Shipping Database\testDatabase.db"
    conn = create_connection(database)
    sql_create_projects_table = """ CREATE TABLE IF NOT EXISTS projects (
                                            name text NOT NULL,
                                            assembly text NOT NULL,
                                            date text
                                        ); """
    sql_create_serialNumber_table = """CREATE TABLE IF NOT EXISTS serialTable (
                                    project_name text NOT NULL,
                                    assembly_name text,
                                    matrix text,
                                    note text,
                                    serialNumber CHAR(25) PRIMARY KEY,
                                    date text,
                                    FOREIGN KEY (project_name) REFERENCES projects (name),
                                    FOREIGN KEY (assembly_name) REFERENCES projects (assembly)
                                );"""
    create_table(conn, sql_create_projects_table)
    create_table(conn, sql_create_serialNumber_table)

def assignInput(listOfInputs):
    test = []
    for i in range(len(listOfInputs)):
        print(listOfInputs[i])
        test.append(input())
    test.append(projectDate)
    return test
    
def enterData():
        database = r"C:\Users\EricChen\Desktop\Shipping Database\testDatabase.db"
        conn = create_connection(database)
        while 1:
            counter = 0
            x = True
            inputList = ["Input Project Name:","Input Assembly:","Input Matrix","Input Note","Scan Serial Number:"]
            data = assignInput(inputList)
            what = create_projects(conn,data[0:2]+data[-1:])
            print("What",what)
            totalList.append(data.copy())
            while x is True:
                counter += 1
                print("Current Number of Boards: "+str(counter))
                print("Scan Again, Press E to Exit, Press N to Input New Project, Press Q to change Note, Press Z to change Matrix,Press D to delete Previous Entry")
                serialNumber = input()
                if serialNumber == 'n':
                    x = False
                elif serialNumber == 'e':
                    for x in totalList:
                        create_serial(conn,x)
                    return 
                elif serialNumber == 'q':
                    z = assignInput([inputList[3],inputList[4]])
                    data[3] = z[0]
                    data[4] = z[1]
                    totalList.append(data.copy())
                    create_serial(conn,data)
                elif serialNumber == 'd':
                    delData=totalList.pop()
                    deleteEntry(conn,delData)
                    counter -= 2
                elif serialNumber == 'z':
                    z = assignInput([inputList[2],inputList[4]])
                    data[2] = z[0]
                    data[4] = z[1]
                    totalList.append(data.copy())
                    create_serial(conn,data)
                else:
                    data[4]=serialNumber
                    totalList.append(data.copy())
                    create_serial(conn,data)
            
def main():
    database = r"C:\Users\EricChen\Desktop\Shipping Database\testDatabase.db"
    conn = create_connection(database)
    createTable()
    if conn is not None:
        menu = ConsoleMenu("Title","Subtitle")
        function_Data = FunctionItem("Enter New Project Data",enterData)
        function_Excel = FunctionItem("Export Data To Excel",exportExcel)
        menu.append_item(function_Data)
        menu.append_item(function_Excel)
        menu.show()
    else:
        print("Error! cannot create the database connection.")

if __name__ == '__main__':
    main()
