from calendar import monthrange
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill

os.chdir(os.path.dirname(__file__))
listOfInputs = []
pendingList = []
rowCount = 0
# Blue 0x0000FF, Yellow 0xFFFF00, Grey 0x808080, Green 0x4CBB17
colorDictionary ={'1':'0000FF','2': 'FFFF00','3':'FFFF00','Z-Completed':'808080','N-Not Ready':'FFFF00','Main WIP':'4CBB17','Mini Done':'4CBB17'}
def getInput(fileName):
    wb = load_workbook(filename=fileName)
    sheet = wb.active
    color = 0
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=22):
        if type(row[0].value) == float:
            color = colorDictionary[str(row[0].value)[0]]
        if row[0].value in colorDictionary.keys():
            color = colorDictionary[row[0].value]
        pc = row[12].value
        li = row[14].value
        aName = row[7].value
        if row[12].value == None:
            pc = ''
        if row[14].value == None:
            li = ''
        if row[7].value == None:
            aName = ''
        if row[1].value == None:
            break
        name = row[1].value+' '+str(pc)+'pcs '+'(LI'+str(li)+') '+aName
        if row[9].value and row[9].value != 'N/A' and row[9].value != 'Pending':
            #listOfInputs format (project name, start date, end date)
            if row[8].value and row[8].value != 'N/A' and row[8].value != 'Pending':
                listOfInputs.append([name + ' Mini',row[8].value,row[8].value,color])
            listOfInputs.append([name+" Main",row[9].value,row[10].value,color])
        elif row[8].value and row[8].value != 'N/A' and row[8].value != 'Pending':
            listOfInputs.append([name + ' Mini',row[8].value,row[8].value,color])
        elif row[9].value=='Pending':
            pendingList.append(name)
        elif row[1].value == None:
            break
        else:
            continue

def addName(projectName,inputCell,sheet,file,w,color):
    global rowCount
    thin = Side(border_style="thin", color="000000")
    if rowCount > 0:
        curCell = sheet.cell(row = rowCount,column = inputCell.column,value = projectName)
        curCell.fill = PatternFill("solid", fgColor=color)
        curCell.border = Border(bottom = thin,top=thin,right=thin)
    else:    
        for x in range(inputCell.row+1,100):
            if isinstance(sheet.cell(row = x,column = inputCell.column).value,str): 
                continue
            if isinstance(sheet.cell(row = x,column = inputCell.column).value,int):
                sheet.insert_rows(x,1)
                curCell = sheet.cell(row = x,column = inputCell.column,value = projectName)
                curCell.fill = PatternFill("solid", fgColor=color)
                for row in sheet.iter_rows(min_row=x-1,max_row=x):
                    for cell in row:
                        cell.border = Border(left=thin,right = thin,bottom = None)
                rowCount = x
                break
            else:
                curCell = sheet.cell(row = x,column = inputCell.column,value = projectName)
                curCell.fill = PatternFill("solid", fgColor=color)
                curCell.border =Border(left=thin,bottom = thin,top=thin,right=thin)
                rowCount = x
                break
    w.save(filename=file)

def parser(file,month,startDay,endDay,projectName,color):
    global rowCount
    wb = load_workbook(filename=file)
    sheet = wb.sheetnames[month-1]
    #exclude weekends by decreasing max col by 2
    for row in wb[sheet].iter_rows(min_row=6,min_col=2,max_col=8):
        for x in row:
            if type(x.value)==int:
                if startDay<=x.value<=endDay :
                    addName(projectName,x,wb[sheet],file,wb,color)
        rowCount = 0
    wb.save(filename=file)

def addToPending(file,pending):
    wb = load_workbook(filename=file)
    sheet = wb['Pending']
    for x in range(len(pending)):
        sheet.cell(row = x+1, column = 1, value = pending[x])
    wb.save(filename=file)
getInput('Test2.xlsx')
# if the end date month and start date month are the same then the month value can be used for both dates
listOfInputs.sort(key=lambda x: x[1])
for x in range(len(listOfInputs)):
    if listOfInputs[x][1].timetuple()[1] == listOfInputs[x][2].timetuple()[1]:
        # parser input format excel file name,month number,day, end day, projectname)
        parser('2022 Calendar.xlsx',listOfInputs[x][1].timetuple()[1],listOfInputs[x][1].timetuple()[2],listOfInputs[x][2].timetuple()[2],listOfInputs[x][0],listOfInputs[x][3])
    else:
        #format is file name,start month,start day,end day,project name)
        endDate = monthrange(listOfInputs[x][1].year,listOfInputs[x][1].timetuple()[1])
        parser('2022 Calendar.xlsx',listOfInputs[x][1].timetuple()[1],listOfInputs[x][1].timetuple()[2],endDate[1],listOfInputs[x][0],listOfInputs[x][3])
        parser('2022 Calendar.xlsx',listOfInputs[x][2].timetuple()[1],1,listOfInputs[x][2].timetuple()[2],listOfInputs[x][0],listOfInputs[x][3])
addToPending('2022 Calendar.xlsx',pendingList)