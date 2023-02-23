from os.path import exists
from os import path
from openpyxl import load_workbook
import os
from collections import defaultdict

x,y,z,s,w,spi,str_list,P_Array,C_Array =[],[],[],[],[],[],[],[],[]
file_list = ["AOI-TOP.txt","AOI-BOT.txt"]
apn = ["APN","Apple PN","Part#","Part #"]
refName = ["Com",'Ref',"REF","ref"]


# Function that creates the folders to store the created files
def createFolders(side):
    # side is an input variable that differentiates between bottom and top side of the board
    # folderPath gets the current directory and adds the side to the directory path 
    folderPath = os.getcwd()+side
    #checks to see if the path exists if it doesn't exist it will make the folders if it does exist it won't do anything
    if not path.exists(folderPath):
        #this is the line that makes the folder
        os.makedirs(folderPath)
# This function will iterate trhough the excel sheet columns and get the column values
def getColumn(cellValue,ar,wbSheet):
    for row in wbSheet.iter_cols(min_col=wbSheet[cellValue].column,max_col=wbSheet[cellValue].column,min_row=int(cellValue[1:]),max_row=200):
        for cell in row:
            ar.append(cell.value)
    return ar
#This function takes an input list and and input string that is the file name and data
def writetotextfile(side,tf):
    #for loop that iterates through the list(line of info) then writes it to the input file
        for d in side:
            #adds a tab after each value in the list
            tf.write(str(d)+"\t")
        #adds a newline onces it exits out of for loop
        tf.write('\n')   

#This function is to check for collected sequential numbers and expand it back out
#EX. 1-5 will become 1,2,3,4,5
def expandArray(testCase):
    #split case will be a list of the min and max range of numbers needed
    splitCase = testCase.split("-")
    for x in range(int(splitCase[0][1:])+1,int(splitCase[1][1:])):
        splitCase.append(splitCase[0][0]+str(x))
    return splitCase

def buildMatrix():
    global str_list,re
    applePN=ref_des=''
    wb = load_workbook(filename = 'bb.xlsx')
    sheet = wb.active
    merged_ranges = sheet.merged_cells.ranges
    if merged_ranges:
        for x in merged_ranges[:]:
            sheet.unmerge_cells(str(x))
    for row in sheet.iter_cols(min_row=1, min_col=1, max_row=100, max_col=50):
        for cell in row:
            if str(cell.value)[0:3] in refName and ref_des == '':
                ref_des = cell.coordinate
            if cell.value in apn and applePN == '':
                applePN = cell.coordinate
    if applePN and ref_des:
        getColumn(applePN,P_Array,sheet)
        getColumn(ref_des,C_Array,sheet)
        h = defaultdict(list)
        for k, v in zip(P_Array, C_Array):
            if v and type(v) == type('string'):
                v = v.replace(' ','')
                v= v.split(',')
                what = [v.extend(expandArray(s)) if "-" in s else s for s in v]
                str_list.extend(what)
                for x in v:
                    h[k].append(x)
        re = dict(h)

def topandbot(sorted,topfile,botfile):
    for side in sorted:
        if "NO" in side[-1]:
            writetotextfile(side[:-1],topfile)
        if "YES" in side[-1]:
            writetotextfile(side[:-1],botfile)
    topfile.close(),botfile.close()

def readCompFile(inputArray):
    for line in inputArray:
        x.append(line.strip().split(','))
    for item in x[5:]:
        y.append(item[0:1]+item[-4:-1]+['T']+item[-1:])
    inputArray.close()
    x.clear()
    y.sort()

def readBomFile():
    wb = load_workbook(filename="bom.xlsx")
    sheet = wb.active
    for row in sheet.iter_rows(min_row = 1,max_row = None,min_col=1,max_col=2):
        z.append([str(row[0].value),str(row[1].value)])
    z.sort()

def createAOI():
    global str_list
    if exists("bb.xlsx"):
       buildMatrix()
    for n in y:
        if n[0] in (item for sublist in z for item in sublist):#loops to check if part number/ref des is in the bom file
            s.append(n)
        if n[0] in str_list: #checks to see if the refdes is in the build matrix
            x.append(n[:-1]+[k for k,v in re.items() if n[0] in v]+n[-1:])
    lst2 = [item[0] for item in s]
    newBomList = []
    for item in z:
        if item[0] in lst2:
            newBomList.append(item)
    for i in range(len(s)):#for loop that combines and formats the comp with the bom
        w.append(s[i][:-1]+newBomList[i][1:2]+s[i][-1:])

    for item in x:# loops through refdes in the build matrix 
        if item in w or None in item or item[0] in lst2:#checks to see if the item is already in w or if its in the combined comp+bom list already
            continue
        else:
            #if its not in those three conditions it will append the item to the final list
            w.extend([item])
    if x:
        multi = open('BoBM.txt','w')
        for v in x:
            writetotextfile(v,multi)

def createFiles(inputFile,outputFile1,outputFile2):
    Top = open(outputFile1,'w')
    Bot = open(outputFile2,'w')
    f = open(inputFile,"r",encoding='utf-8')
    readCompFile(f)
    readBomFile()
    createAOI()
    topandbot(w,Top,Bot)

def changExtension():
    for file in file_list:
        if exists(file):
            path = os.getcwd()
            b = os.path.basename(path)
            if not exists(b[:-23]+file[-7:-4]+'.aoi'):
                os.rename(file,b[:-23]+file[-7:-4]+'.aoi')

if exists("comp.txt") and exists("bom.xlsx"):
    createFiles("comp.txt","AOI-TOP.txt","AOI-BOT.txt")
    changExtension()
    createFolders("\TOP")
    createFolders("\BOT")
else:
    print("No Comp.txt and BOM.xlsx found")
