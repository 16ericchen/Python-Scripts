from openpyxl import load_workbook,Workbook
import os
from os import path
os.chdir(path.dirname(__file__))

listofBOMS,deleted = [],[]
newRev,preRev,newQty,added,itemCat = {},{},{},{},{}

def firstParser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, min_col=1):
        populateDict(preRev,row,None)

def secondParser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, min_col=1):
        deleted.append((row[0].value,row[1].value,row[3].value))
        if (row[0].value,row[1].value,row[3].value) not in preRev.keys():
            populateDict(added,row,None)
        if (row[0].value,row[1].value,row[3].value) in preRev.keys():
            if row[2].value != preRev[(row[0].value,row[1].value,row[3].value)][2]:
                d = (row[2].value) - int(preRev[(row[0].value,row[1].value,row[3].value)][2])
                if row[3].value != preRev[(row[0].value,row[1].value,row[3].value)][3]:
                    populateDict(itemCat,row,d)
                else:
                    populateDict(newQty,row,d)
            else:
                continue

def populateDict(inputDict,row,diff):
    inputDict[row[0].value,row[1].value,row[3].value] = []
    if diff:
        if diff<0:
            print('increased by: '+str(diff))
            inputDict[row[0].value,row[1].value,row[3].value].append('increased by: '+str(abs(diff)))
        else:
            print('decreased by: '+ str(abs(diff)))
            inputDict[row[0].value,row[1].value,row[3].value].append('decreased by: '+ str(abs(diff)))
    for item in row:
        inputDict[row[0].value,row[1].value,row[3].value].append(item.value)

def writeData(title,titleRow,data,sheet):
    sheet.cell(row = titleRow, column = 1, value = title)
    rowN = 0
    if len(data)>0:
        for x in data:
            rowN +=1
            colN = 1
            for y in data[x]:
                log.cell(row=rowN+titleRow, column = colN+1,value = y)
                colN+=1

def writeDataDeleted(title,titleRow,data,sheet):
    sheet.cell(row = titleRow, column = 1, value = title)
    rowN = 0
    if len(data)>0:
        for x in data:
            rowN +=1
            colN = 1
            for y in preRev[x]:
                log.cell(row=rowN+titleRow, column = colN+1,value = y)
                colN+=1

for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')) and x !='BOM_Comparison.xlsx':
        listofBOMS.append(x)
wb = Workbook()
log = wb.active
log.title = 'BOMComparison'
firstParser(listofBOMS[0])
secondParser(listofBOMS[1])
difference = preRev.keys()-deleted
writeData('New Parts',1,added,log)
writeData('Updated Quantity',len(added)+3,newQty,log)
writeData('Changed Item Cat',len(added)+len(newQty)+5,itemCat,log)
writeDataDeleted('Deleted Parts',len(added)+len(newQty)+len(itemCat)+7,difference,log)

wb.save(filename='BOM_Comparison.xlsx')              