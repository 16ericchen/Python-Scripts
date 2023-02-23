import os
import shutil
from openpyxl import load_workbook
def spi():
    desktop = os.path.join(os.environ['USERPROFILE'],'Desktop','SPI') 
    projectListPath = r"K:\Testing\SPI\SPI Program List.xlsx"
    projectInfo = ''
    wb = load_workbook(projectListPath)
    print("Input Board Number:")
    boardNumber = input()
    ws = wb.active
    for cell in ws['A']:
        if boardNumber in cell.value:
            projectInfo = cell.value
    if projectInfo == '':
        print("Board Number Not Found")
        spi()
    projectInfo = projectInfo.replace(',','')
    if not os.path.exists(desktop):
        os.makedirs(desktop)
    spiNTDPath = os.path.join(desktop,projectInfo)
    if not os.path.exists(spiNTDPath):
        os.makedirs(spiNTDPath)
    projectLetter = projectInfo[0]
    projectInfo = projectInfo.replace(',','')
    test = projectInfo.split(' ')
    test = [i for i in test if i]
    projectName = ' '.join(test[0:-2])
    if 'board' in projectInfo:
        simpleName = projectInfo.split('board')[0]
    if 'board' not in projectInfo and simpleName == '':
        simpleName = projectInfo.split('Proto')[0]
    projectNumber = test[0]
    projectDate = test[-1]
    projectFolderPath = "K:\Projects"+"\\"+projectLetter+"\\"+projectNumber
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        spi()
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if x.replace(' ','').lower() in projectName.replace(' ','').lower() or simpleName.replace(' ','').lower() in x.replace(' ','').lower():
            fullName =projectFolderPath+ "\\"+x
            if simpleName.replace(' ','').lower() == x.replace(' ','').lower():
                break
    projectFolderPath = fullName
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        spi()
    folders = next(os.walk(projectFolderPath))[1]
    for x in folders:
        if (x[:-11] in projectName or projectDate==folders[0][-10:]) and x[:-11] != '':
            projectFolderPath = projectFolderPath +"\\"+x
            break
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        spi()
    if not os.path.exists(projectFolderPath):
        print('Folder Path Does Not Exists: ',projectFolderPath)
        spi()
    compPath = projectFolderPath + '\\'+'Gerber'+'\\'+'comp.txt'
    if not os.path.exists(compPath):
        print("Comp File Not Found")
        spi()
    if os.path.exists(compPath):
        shutil.copy(compPath,spiNTDPath)
        print("COMP File was Moved")
    designPath = projectFolderPath + '\\'+'Download'+'\\'+boardNumber+'.tgz'
    if not os.path.exists(designPath):
        print("TGZ File Not Found")
        spi()
    if os.path.exists(designPath):
        shutil.copy(designPath,spiNTDPath)
        print("TGZ File was Moved")
    spi()
spi()
