import os
import shutil
from openpyxl import load_workbook
projectListPath = r"C:\Users\EricChen\Desktop\SPI Need To Do\SPI Program List Test.xlsx"
wb = load_workbook(projectListPath)
print("Input Board Number:")
boardNumber = input()
ws = wb.active
for cell in ws['A']:
    if boardNumber in cell.value:
        projectInfo = cell.value
        print(projectInfo)
folderPath = r"C:\Users\EricChen\Desktop\AOI"+'\\'+projectInfo
if not os.path.exists(folderPath):
    os.makedirs(folderPath)
projectLetter = projectInfo[0]
projectInfo = projectInfo.replace(',','')
test = projectInfo.split(' ')
projectName = ' '.join(test[0:-2])
projectNumber = test[0]
projectDate = test[-1]
projectFolderPath = "K:\Projects"+"\\"+projectLetter+"\\"+projectNumber
if os.path.exists(projectFolderPath):
    print("projectFolderPath exists")
folders = next(os.walk(projectFolderPath))[1]
for x in folders:
    if x in projectName:
        projectFolderPath =projectFolderPath+ "\\"+x+"\\"+projectName+" "+projectDate
if os.path.exists(projectFolderPath):
    print('projectFolderPath exists2',projectFolderPath)
compPath = projectFolderPath + '\\'+'Gerber'+'\\'+'comp.txt'
if not os.path.exists(compPath):
    print("Comp File Not Found")
if os.path.exists(compPath):
    shutil.copy(compPath,folderPath)
bomPath = projectFolderPath + '\\'+'BOM_CAD'
for x in os.listdir(bomPath):
    if 'bom' in x.lower():
        shutil.copy(bomPath+"\\"+x,folderPath)
if not os.path.exists(bomPath):
    print("BOM File Not Found")
if os.path.exists(compPath):
    shutil.copy(compPath,folderPath)
buildMatrixPath = projectFolderPath +'\\'+'Information'
for x in os.listdir(buildMatrixPath):
    if x.endswith(('.xlsx','xls')):
        shutil.copy(buildMatrixPath+"\\"+x,folderPath)

