from email.encoders import encode_noop
from openpyxl import load_workbook
import sqlite3
from sqlite3 import Error

wb = load_workbook(filename = 'SHIPLOG 03-2022.xlsx')
sheets = wb.sheetnames

def writetotextfile(side,tf):
    for d in side:
        tf.write(str(d)+"\t")
    tf.write('\n')  
    tf.close()

def getData(ws,projDate):
    # for loop to get the max amount of columns 
    info,rowInfo = [],[]
    counter = 0
    for row in ws.iter_cols(min_col=0,max_col=None,min_row=2,max_row=2,values_only=True,):
        if counter == 5:
            counter = 0
        if counter == 0 and not row[0]:
            break
        else:
            info.append(row[0])
            counter+=1
    # for loop to retrieve data
    for i in range(1,len(info)+1,5):
        data = []
        data.append(info[i-1])
        # max rows loop
        for row in ws.iter_rows(min_col=i+2,max_col=i+2,min_row=2,values_only=True):
            if row[0] is not None:
                counter += 1
        # loops through rows to get data
        for rows in ws.iter_rows(min_col= i+1,max_col =i+4,min_row=2,max_row=counter+1,values_only=True):
            data.append(rows)
        counter = 0
        rowInfo.append(data)
    enterData(rowInfo,projDate)

def enterData(data,projDate):
    database = r"C:\Users\EricChen\Desktop\Shipping Database\shippingDatabase.db"
    conn = create_connection(database)
    for x in range(len(data)):
        projectName = data[x][0]
        projectAssembly = 'Null '
        for y in data[x][1:]:
            if y[0] is not None:
                projectAssembly = y[0]
            serialNumber = y[1]
            projectMatrix = y[2]
            projectNote = y[3]
            projectDate = projDate
            task_1 = (projectName,projectAssembly,projectDate)
            create_projects(conn,task_1)
            task_2 = (projectName,projectAssembly,serialNumber,projectMatrix,projectNote,projectDate)
            create_serial(conn,task_2)

def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except Error as e:
        print(e)
    return conn

def create_table(conn, create_table_sql):
    try:
        c = conn.cursor()
        c.execute(create_table_sql)
    except Error as e:
        print(e)

def create_projects(conn, projects):
    sql = ''' INSERT INTO projects(name,assembly,date)
              VALUES(?,?,?) '''
    cur = conn.cursor()
    cur.execute("SELECT * FROM projects WHERE (name,assembly,date) = (?,?,?)",(projects[0],projects[1],projects[2],))
    x = cur.fetchall()
    if not x:
        cur.execute(sql, projects)
    conn.commit()
    return cur.lastrowid

def create_serial(conn, projects):
    f = open('duplicate.txt','a',encoding='utf-8')
    sql = ''' INSERT INTO serialTable(project_name,assembly_name,serialNumber,matrix,note,date)
              VALUES(?,?,?,?,?,?) '''
    cur = conn.cursor()
    cur.execute("SELECT * FROM serialTable WHERE serialNumber=?", (projects[2], ))
    x =cur.fetchone()
    if x:
        writetotextfile(projects,f)
    else:
        cur.execute(sql, projects)
    conn.commit()
    return cur.lastrowid

def createTable():
    database = r"C:\Users\EricChen\Desktop\Shipping Database\shippingDatabase.db"
    conn = create_connection(database)
    sql_create_projects_table = """ CREATE TABLE IF NOT EXISTS projects (
                                            name text,
                                            assembly text NOT NULL,
                                            date text
                                        ); """
    sql_create_serialNumber_table = """CREATE TABLE IF NOT EXISTS serialTable (
                                    project_name text NOT NULL,
                                    assembly_name text,
                                    serialNumber CHAR(25) PRIMARY KEY,
                                    matrix text,
                                    note text,
                                    date text,
                                    FOREIGN KEY (project_name) REFERENCES projects (name),
                                    FOREIGN KEY (assembly_name) REFERENCES projects (assembly)
                                );"""
    create_table(conn, sql_create_projects_table)
    create_table(conn, sql_create_serialNumber_table)
def main():
    database = r"C:\Users\EricChen\Desktop\Shipping Database\shippingDatabase.db"
    conn = create_connection(database)
    createTable()
    if conn is not None:
        for workSheets in sheets:
            getData(wb[workSheets],workSheets)
    else:
        print("Error! cannot create the database connection.")

if __name__ == '__main__':
    main()

