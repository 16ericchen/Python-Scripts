from playwright.async_api import async_playwright
import time,re,asyncio,os
from searchDB_WS import searchData
from openpyxl.styles import Font,PatternFill
from yageoTable import table
from openpyxl import load_workbook,Workbook
import math
start_time = time.time()
# os.chdir(os.path.dirname(__file__))
# $env:PLAYWRIGHT_BROWSERS_PATH="0" playwright install firefox pyinstaller --onefile newExcelParser.py
yageoTable,yageoAPN,apnDict,mfrDict,final,notFound = {},{},{},{},{},{}
testList = ['Mouser Electronics','Digi-Key','Arrow Electronics','Future Electronics','Verical','Newark','Master','Rochester','Master Electronics']
octoPathList = ['Mouser','Digi-Key','Arrow Electronics','Future Electronics','Verical','Newark','Master','Rochester','Master Electronics']
cyntecDict = {}
addToDB = []
titles = ["APN",'Quantity','Price Break',"Distributor","MFR","Stock","MPN","Per","Per Price"]
def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    if sheet == None:
        print('Excel Format Cannot Be Read')
        print('Change File Type From Strict Open XML Spreadsheet To Regular XLSX Format')
        print('Then Restart the Program')
        input()
    for row in sheet.iter_rows(min_col=3,min_row=1,max_row=sheet.max_row):
        apn = sheet.cell(row=row[0].row,column=1).value
        quantity = sheet.cell(row=row[0].row,column=2).value
        i = 0
        if row[i].value:
            apnDict[apn] = [quantity]
        while row[i].value:
            manufacturer = row[i].value.lower()
            if 'cyntec' in manufacturer or 'murata' in manufacturer:
                cyntecDict[apn] = [quantity,row[i].value,row[i+1].value]
                if apn in yageoAPN.keys():
                    del yageoAPN[apn]
                if apn in apnDict.keys():
                    del apnDict[apn]
                break
            elif 'yageo' in manufacturer:
                if apn in yageoAPN.keys():
                    if row[i+1].value not in yageoAPN[apn]:
                        yageoAPN[apn].append(row[i+1].value)
                    i+=2
                else:
                    yageoAPN[apn] = [row[i+1].value]
                    i+=2
            else:
                if row[i+1].value not in apnDict[apn]:
                    apnDict[apn].append(row[i+1].value)
                i+=2
    if yageoAPN:
        for x in yageoAPN:
            for y in yageoAPN[x]:
                if searchData(y):
                    apnDict[x].append(searchData(y))
                else:
                    if x in yageoTable.keys():
                        yageoTable[x].append(y)
                        addToDB.append(y)
                    else:
                        addToDB.append(y)
                        yageoTable[x] = [y]
    if yageoTable:
        table(addToDB)
        for x in yageoTable:
            for y in yageoTable[x]:
                    if searchData(y):
                        apnDict[x].append(searchData(y))
    counter = 0
    for x in apnDict:
        for y in range(1,len(apnDict[x])):
            functionSearch = asyncio.run(findChips(int(apnDict[x][0]),str(apnDict[x][y])))
            if functionSearch[1] == True:
                final[x] = [apnDict[x][0],'']+list(functionSearch[0].values())[0]
                break
            functionSearch = asyncio.run(octoPath(functionSearch[2],str(apnDict[x][y])))
            if functionSearch[1] == True:
                final[x] = [apnDict[x][0],'']+list(functionSearch[0].values())[0]
                break
            functionSearch = asyncio.run(oemTrades(functionSearch[2],str(apnDict[x][y])))
            if functionSearch[1] == True:
                final[x] = [apnDict[x][0],'']+list(functionSearch[0].values())[0]
                break
        if x not in final.keys():
            notFound[x] = ["Could Not Find Item"]
        counter+=1
        print(str(counter)+'/'+str(len(apnDict))) 
    createXL()
def createXL():
    wx = Workbook()
    log = wx.active

    cellFont = Font(bold=True)
    for x in range(len(titles)):
        log.cell(row = 1,column = 1+x,value=titles[x])
    rnum = 2
    for x in final:
        curPos = [rnum,6]
        curMin = math.inf
        log.cell(row = rnum,column=1,value = x)
        for y in range(len(final[x])):
            bCell=log.cell(row=rnum,column=y+2,value=final[x][y])
            if y>=6 and y%2 == 0:
                test = final[x][y+1].replace('$','')
                if int(final[x][y])<= final[x][0]:
                    curMin = min(curMin,final[x][0]*float(test))
                    if curMin == final[x][0]*float(test):
                        curPos = [rnum,y+2]
                    bCell.font=cellFont
                else:
                    curMin = min(curMin,int(final[x][y])*float(test))
                    if curMin == int(final[x][y])*float(test):
                        curPos = [rnum,y+2]
                    bCell.font=cellFont
        log.cell(row=rnum,column=3,value=curMin)
        rnum+=2
        if curPos:    
            breakCell = log.cell(row=curPos[0],column=curPos[1])
            breakCell.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid")


# def createXL():
#     wx = Workbook()
#     log = wx.active
#     cellFont = Font(bold=True)
#     for x in range(len(titles)):
#         log.cell(row = 1,column = 1+x,value=titles[x])
#     rnum = 2
#     for x in final:
#         curMin = [rnum,6]
#         log.cell(row = rnum,column=1,value = x)
#         for y in range(len(final[x])):
#             bCell=log.cell(row=rnum,column=y+2,value=final[x][y])
#             if y>=5 and y%2 == 1:
#                 print(type(final[x][0]),final[x][y])
#                 if final[x][0]>=int(final[x][y]):
#                     curMin = [rnum,y+3]
#                     bCell.font=cellFont
#                 else:
#                     print('here')
#                     breakCell = log.cell(row=curMin[0],column=curMin[1],value=final[x][curMin[1]])
#                     breakCell.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid")
#                     bCell.font=cellFont
#                     break
#         rnum+=2
    for j in notFound:
        log.cell(row = rnum,column=1,value = j)
        for y in range(len(notFound[j])):
            log.cell(row=rnum,column=y+2,value=notFound[j][y])
        rnum+=1
    for j in cyntecDict:
        log.cell(row = rnum,column=1,value = j)
        for y in range(len(cyntecDict[j])):
            log.cell(row=rnum,column=y+2,value=cyntecDict[j][y])
        rnum+=1
    wx.save(filename='bom.xlsx')

# Octopath
async def octoPath(quantity,pn):
    async with async_playwright() as pw:
        browser = await pw.firefox.launch()
        browser2 = await browser.new_context(ignore_https_errors=True)
        page = await browser2.new_page()
        result = {}
        await page.goto("https://octopart.com/search?q="+pn+"&currency=USD&specs=0")
        noResult = await page.query_selector('//div[contains(@class, "no-results-found")]')
        if noResult:
            await browser.close()
            return (result,False,quantity)
        distributor = await page.query_selector('//div[contains(@class, "prices-view")]')
        if not distributor:
            return (result,False,quantity)
        distributorTest = await distributor.query_selector_all('//div[contains(@class, "part")]')
        for x in range(len(distributorTest)):
            header = await distributorTest[x].query_selector('//div[contains(@class, "header")]')
            left = await header.query_selector('//div[contains(@class, "left")]')
            middle = await left.query_selector('//div[contains(@class, "middle")]')
            mnf = await middle.query_selector('//div[contains(@class, "manufacturer-name-and-possible-tooltip")]')
            mpn = await middle.query_selector('//div[contains(@class, "mpn")]')
            footer = await distributorTest[x].query_selector('//div[contains(@class, "footer")]')
            button = await footer.query_selector('//button[contains(@type, "button")]')
            if button:  
                if await button.inner_text() == "Show All":
                    await button.click()
            for y in octoPathList:
                if quantity <= 0 :
                    await browser.close()
                    return (result,True)
                listOfDistributor = await distributorTest[x].query_selector_all('//td[contains(., "'+y+'")]')
                if listOfDistributor:
                    getResult = await asyncio.create_task(opPriority(listOfDistributor,quantity,await mnf.inner_text(),await mpn.inner_text()))
                    quantity = getResult[1]
                    result.update(getResult[0])
                else:
                    continue
        await browser.close()
        return (result,False,quantity)

async def opPriority(items,quantity,mnf,mpn):
    columnItems = {}
    for x in items:
        parent = await x.query_selector('xpath=..')
        columns = await parent.query_selector_all('td')
        stock = await columns[3].inner_text()
        stock = int(stock.replace(',',''))
        if quantity<=0:
            return (columnItems,quantity)
        else:
            quantity -= int(stock)
            testName = await columns[1].inner_text()
            distributorName = testName.split('\n',1)[0]
            columnItems[distributorName+'-'+await columns[3].inner_text()] = [distributorName,mnf,await columns[3].inner_text(),mpn,'1',await columns[7].inner_text(),'10',await columns[8].inner_text(),'100',await columns[9].inner_text(),'1000',await columns[10].inner_text(),'10000',await columns[11].inner_text()]
    return (columnItems,quantity)

# OEMTrades
async def oemTrades(quantity,pn):
    async with async_playwright() as pw:
        browser = await pw.firefox.launch()
        browser2 = await browser.new_context(ignore_https_errors=True)
        page = await browser2.new_page()
        output ={}
        await page.goto("https://www.oemstrade.com/search/"+pn)
        noResult = await page.query_selector('//div[contains(@class, "no-results")]')
        if noResult:
            await browser.close()
            return (output,False,quantity)
        for x in testList:
            if quantity <= 0:
                await browser.close()
                return (output,True)
            distributor = await page.query_selector('//div[contains(@data-distributorname, "' +x+'")]')
            if distributor:
                if 'CIC' in await distributor.inner_text():
                    break 
                getResult = await asyncio.create_task(otPriority(distributor,quantity))
                quantity = getResult[1]
                output.update(getResult[0])
            else:
                continue
        await browser.close()
        return (output,False,quantity)

async def otPriority(distributor,quantity):
    if distributor:
        result = {}
        distributorName = await distributor.query_selector('.distributor-title')
        subtract = await distributorName.query_selector('span')
        z = await distributorName.inner_text()
        w = await subtract.inner_text()
        # distributor
        z = z.replace(w,'')
        getInfo = await distributor.query_selector('tbody')
        rows = await getInfo.query_selector_all('tr')
        for x in range(len(rows)):
            # instock value
            stock = await rows[x].query_selector('.td-stock')
            stockk = await stock.inner_text()
            stockk = re.sub("[^0-9]", "", stockk.replace("\n", ""))
            if int(stockk) > 0:
                prices = await rows[x].query_selector('.td-price')
                click = await prices.query_selector(".show-more")
                if click:
                    await click.click()
                # prices
                mnf = await rows[x].query_selector('.td-part-number')
                mnfValue = await mnf.query_selector('a')
                manufacturer = await rows[x].query_selector('.td-distributor-name')
                # prices
                priceList = await prices.inner_text()
                priceList = priceList.split('\n')
                quantity -= int(stockk)
                if quantity<=0:
                    result[z.strip()+"-"+stockk] = [z.strip(),await manufacturer.inner_text(),stockk,await mnfValue.inner_text()]+priceList
                    return (result,quantity)
                else:
                    result[z.strip()+"-"+stockk] = [z.strip(),await manufacturer.inner_text(),stockk,await mnfValue.inner_text()]+priceList
        return (result,quantity)

# FindChips
async def findChips(quantity,pn):
    async with async_playwright() as pw:
        output = {}
        browser = await pw.firefox.launch()
        browser2 = await browser.new_context(ignore_https_errors=True)
        page = await browser2.new_page()
        if "/" in pn:
            pn = pn.replace('/',"%2F")
        await page.goto("https://www.findchips.com/search/"+pn,timeout=0)
        noResult = await page.query_selector('//p[contains(@class, "no-results")]')
        if noResult:
            await browser.close()
            return (output,False,quantity)
        for x in testList:
            if quantity <= 0:
                await browser.close()
                return (output,True)
            distributor = await page.query_selector('//div[contains(@data-distributor_name, "' +x+'") and contains(@class, "distributor-results")]')
            if distributor:
                if 'CIC' in await distributor.inner_text():
                    break
                getResult = await asyncio.create_task(fcPriority(distributor,quantity))
                quantity = getResult[1]
                output.update(getResult[0])
            else:
                continue
        await browser.close()
        return (output,False,quantity)

async def fcPriority(distributor,quantity):
    if distributor:
        result = {}
        test = await distributor.get_attribute('data-distributor_name')
        getInfo = await distributor.query_selector('tbody')
        rows = await getInfo.query_selector_all('tr')
        for x in range(len(rows)):
            # instock value
            stockk =await rows[x].get_attribute('data-instock')
            if int(stockk) > 0:
                prices = await rows[x].query_selector('.td-price')
                click = await prices.query_selector(".hyperlink")
                if click:
                    await click.click()
                # prices
                mnf = await getInfo.query_selector('.td-mfg')
                # part number
                pn = await rows[x].get_attribute('data-mfrpartnumber')
                priceList = await prices.inner_text()
                priceList = priceList.split('\n')
                quantity -= int(stockk)
                if quantity<=0:
                    result[test+"-"+stockk] = [test,await mnf.inner_text(),stockk,pn]+priceList
                    return (result,quantity)
                else:
                    result[test+"-"+stockk] = [test,await mnf.inner_text(),stockk,pn]+priceList
        return (result,quantity)

if __name__ == '__main__':
    flag = False
    for x in os.listdir(os.getcwd()):
        if x.endswith(('.xlsx')):
            print(x)
            flag = True
            parser(x)
        if x.endswith(('.xls')):
            print('Change File Type from XLS to XLSX Format')
            print('Then Restart the Program')
            input()
    if flag == False:
        print("No Excel File Found")
        input()
    print("Process finished --- %s seconds ---" % (time.time() - start_time))
    input()




