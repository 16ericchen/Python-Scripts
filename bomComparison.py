from openpyxl import load_workbook,Workbook
import os
from os.path import exists
os.chdir(os.path.dirname(__file__))
# def parser(file):
#     print('parsed'+': '+file)
#     wb = load_workbook(filename=file)
#     sheet = wb.active

def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, min_col=1):
        if row:
            component[row[0].value] = [row[1].value]
        else:
            break

def compare(list,count):
    wb = load_workbook(filename=list)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, min_col=1):
        if row:
            #row[0].value == location ex. R1714
            #row[1].value == part number ex. 101S0075
            if row[0].value in component.keys():
                if component[row[0].value] == row[1].value:
                    continue
                else:
                    #add the part number to the dictionary corresponding to the same location
                    if row[0].value in comparisonResult.keys():
                        comparisonResult[row[0].value] = comparisonResult[row[0].value] + [row[1].value]
                    else:
                        comparisonResult[row[0].value] = component[row[0].value] + [row[1].value]
                    print('Same location But different Part Number')
            else:
                test = ['No Stuff']*count
                if row[0].value in comparisonResult.keys():
                    comparisonResult[row[0].value] = comparisonResult[row[0].value] + [row[1].value]
                else:
                    comparisonResult[row[0].value] = test + [row[1].value]
                print("Location Not in Bom")
        else:
            break


def exportResult():
    wb = Workbook()
    log = wb.active
    log.title = 'BOMComparison'
    wb.save(filename='BOMComparison')
listofBOMS = []
component = {}
comparisonResult = {}
for x in os.listdir(os.getcwd()):
    if x.endswith(('.xlsx')):
        listofBOMS.append(x)
parser(listofBOMS[0])
for i in range(1,len(listofBOMS)-1):
    compare(listofBOMS[i],i)


        