import os
from os import path
from openpyxl import load_workbook,Workbook
def parser(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    if sheet == None:
        print('Excel Format Cannot Be Read')
        print('Change File Type From Strict Open XML Spreadsheet To Regular XLSX Format')
        print('Then Restart the Program')
        input()
    dictInfo = {}
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=6):
        #0=part number,1=description,2=qty,3=ref des,4=level,5 = item Cat
        if row[0].value == None:
            continue
        if row[3].value and 'PCB' not in row[3].value and '1' in str(row[4].value) and row[5].value == 'L':
            refTotal = row[3].value.split(',')
            for x in refTotal:
                #location = [APN,Description]
                dictInfo[x] = [row[0].value,row[1].value]
    createXL(dictInfo)
def createXL(data):
    wx = Workbook()
    log = wx.active
    rnum = 1
    for x,y in data.items():
        log.cell(row=rnum,column=1,value= x)
        log.cell(row=rnum,column=2,value= y[0])
        log.cell(row=rnum,column=3,value= y[1])
        rnum+=1
    wx.save(filename='bom.xlsx')

    # while cell
flag = False
for x in os.listdir(os.getcwd()):
    print(x)
    if x.endswith(('.xlsx')):
        print(x)
        flag = True
        parser(x)
    if x.endswith(('.xls')):
        print('Change File Type from XLS to XLSX Format')
        print('Then Restart the Program')
        input()
if flag == False:
    print("No Excel File Found")
    input()